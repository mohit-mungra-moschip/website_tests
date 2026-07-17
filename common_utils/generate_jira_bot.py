"""
Jira Bug Reporter — Automated Bug Creation from Excel Test Results
==================================================================

Purpose:
    This script scans an Excel workbook's "Test Details" sheet for test cases
    marked as FAIL, creates a corresponding Bug issue in Jira for each failure,
    assigns it to the configured assignee, adds it to the active sprint, and
    writes the Jira issue hyperlink back into a "JiraID" column in the sheet.

    It also reads the AI Short Summary and Suggested Fix columns and includes
    them in the Jira description for richer context.

    Optionally, it patches the HTML dashboard (5th argument) so that each
    failed test card shows a clickable Jira link directly in the browser.

    Handles merged-cell test cases by aggregating Steps, Expected Result, and
    Actual Result across all rows that share the same Title before creating a
    single Jira issue.

Usage:
    python 6_generate_jira_bot.py <PROJECT_KEY> <BOARD_NAME> <SPRINT_NAME> <EXCEL_FILE> [HTML_REPORT]
    python common_utils/utils/ai_agents/scripts/6_generate_jira_bot.py SCRUM "SCRUM board" "SCRUM Sprint 0" "/path/to/report.xlsx" "/path/to/report.html"

Arguments:
    PROJECT_KEY   Jira project key (e.g., QA, DEV, TEST)
    BOARD_NAME    Exact name of the Jira board containing the sprint
    SPRINT_NAME   Exact name of the active sprint to add issues to
    EXCEL_FILE    Path to the Excel workbook (.xlsx)
    HTML_REPORT   (Optional) Path to the HTML dashboard to patch with Jira links

Project structure:
    PTAF-Core/
    ├── config.py
    └── common_utils/utils/ai_agents/scripts/
        └── 6_generate_jira_bot.py

Configuration:
    All credentials and default settings are managed in config.py at the
    project root. Update that file before running the script.

Dependencies:
    pip install openpyxl jira
"""

import re
import sys
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Side, Border
from jira import JIRA

# Directory depth: PTAF-Core/common_utils/
PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import config


# ==========================================
# HELPER FUNCTIONS
# ==========================================

def get_effective_cell_value(sheet, row, col):
    """
    Return the effective value of a cell.

    If the cell belongs to a merged range, returns the value of the
    top-left (anchor) cell of that range instead of None.
    """
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value


def are_cells_merged_together(sheet, row1, row2, col):
    """
    Return True if the cells at (row1, col) and (row2, col) belong to the same
    merged cell range in the worksheet.
    """
    coord1 = sheet.cell(row=row1, column=col).coordinate
    coord2 = sheet.cell(row=row2, column=col).coordinate
    for merged_range in sheet.merged_cells.ranges:
        if coord1 in merged_range and coord2 in merged_range:
            return True
    return False


def find_column_by_name(sheet, column_name):
    """
    Scan the worksheet and return (row_index, col_index) of the first cell
    whose value matches *column_name*.  Returns (None, None) if not found.
    """
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == column_name:
                return row, col
    return None, None


def build_border_from_cell(cell):
    """
    Derive a Border style from an existing cell's border, or return None
    if the cell has no border applied.
    """
    if cell.border and cell.border.left.style:
        return Border(
            left=Side(style=cell.border.left.style),
            right=Side(style=cell.border.right.style),
            top=Side(style=cell.border.top.style),
            bottom=Side(style=cell.border.bottom.style),
        )
    return None


def ensure_jira_id_column(sheet, header_row):
    """
    Return the column index for the 'JiraID' column.

    If the column does not yet exist, it is created immediately after the
    last non-empty header column and formatted to match the surrounding cells.
    """
    _, jira_id_col_index = find_column_by_name(sheet, config.COLUMN_JIRA_ID)

    if jira_id_col_index is not None:
        return jira_id_col_index

    last_used_col = max(
        (col for col in range(1, sheet.max_column + 1)
         if sheet.cell(row=header_row, column=col).value not in (None, "")),
        default=0,
    )
    jira_id_col_index = last_used_col + 1

    header_cell = sheet.cell(row=header_row, column=jira_id_col_index)
    header_cell.value = config.COLUMN_JIRA_ID
    header_cell.font = Font(bold=True)

    column_letter = header_cell.column_letter
    sheet.column_dimensions[column_letter].width = config.JIRA_ID_COLUMN_WIDTH

    border_style = build_border_from_cell(sheet.cell(row=header_row, column=1))
    if border_style:
        header_cell.border = border_style
        for r in range(header_row + 1, sheet.max_row + 1):
            sheet.cell(row=r, column=jira_id_col_index).border = border_style

    return jira_id_col_index


def ensure_create_jira_column(sheet, header_row):
    """
    Return the column index for the 'CreateJira' column.

    If the column does not yet exist, it is created immediately after the
    last non-empty header column and formatted to match the surrounding cells.
    """
    _, create_jira_col_index = find_column_by_name(sheet, config.COLUMN_CREATE_JIRA)

    if create_jira_col_index is not None:
        return create_jira_col_index

    last_used_col = max(
        (col for col in range(1, sheet.max_column + 1)
         if sheet.cell(row=header_row, column=col).value not in (None, "")),
        default=0,
    )
    create_jira_col_index = last_used_col + 1

    header_cell = sheet.cell(row=header_row, column=create_jira_col_index)
    header_cell.value = config.COLUMN_CREATE_JIRA
    header_cell.font = Font(bold=True)

    column_letter = header_cell.column_letter
    sheet.column_dimensions[column_letter].width = config.CREATE_JIRA_COLUMN_WIDTH

    border_style = build_border_from_cell(sheet.cell(row=header_row, column=1))
    if border_style:
        header_cell.border = border_style
        for r in range(header_row + 1, sheet.max_row + 1):
            sheet.cell(row=r, column=create_jira_col_index).border = border_style

    return create_jira_col_index



def collect_merged_rows(
    sheet, start_row, title_col, steps_col, expected_col,
    description_col, failure_reason_col, ai_summary_col, ai_fix_col,
    testcase_title
):
    """
    Starting from *start_row*, walk forward and collect consecutive rows that
    share the same *testcase_title* and are physically merged in the Excel sheet.

    Returns:
        merged_rows (list[int]): row indices belonging to this test case.
        steps (str):             aggregated steps text.
        expected_output (str):   aggregated expected result text.
        description (str):       aggregated description text.
        failure_reason (str):    aggregated failure reason text.
        ai_summary (str):        AI short summary text.
        ai_fix (str):            AI suggested fix text.
    """
    steps          = str(sheet.cell(row=start_row, column=steps_col).value or "")
    expected_output= str(sheet.cell(row=start_row, column=expected_col).value or "")
    description    = str(sheet.cell(row=start_row, column=description_col).value or "")
    failure_reason = str(sheet.cell(row=start_row, column=failure_reason_col).value or "")
    ai_summary     = str(sheet.cell(row=start_row, column=ai_summary_col).value or "") if ai_summary_col else ""
    ai_fix         = str(sheet.cell(row=start_row, column=ai_fix_col).value or "")     if ai_fix_col     else ""
    merged_rows    = [start_row]

    if not testcase_title or str(testcase_title).strip().upper() == "N/A":
        return merged_rows, steps, expected_output, description, failure_reason, ai_summary, ai_fix

    next_row = start_row + 1
    while next_row <= sheet.max_row:
        if get_effective_cell_value(sheet, next_row, title_col) == testcase_title:
            if are_cells_merged_together(sheet, start_row, next_row, title_col):
                add_steps          = sheet.cell(row=next_row, column=steps_col).value or ""
                add_expected       = sheet.cell(row=next_row, column=expected_col).value or ""
                add_description    = sheet.cell(row=next_row, column=description_col).value or ""
                add_failure_reason = sheet.cell(row=next_row, column=failure_reason_col).value or ""

                if add_steps:          steps          += "\n" + str(add_steps)
                if add_expected:       expected_output += "\n" + str(add_expected)
                if add_description:    description    += "\n" + str(add_description)
                if add_failure_reason: failure_reason += "\n" + str(add_failure_reason)

                merged_rows.append(next_row)
                next_row += 1
                continue
        break

    return merged_rows, steps, expected_output, description, failure_reason, ai_summary, ai_fix


def build_jira_description(description, steps, expected_output, failure_reason, ai_summary, ai_fix):
    """
    Build a well-formatted Jira description using wiki markup.

    Sections are separated by blank lines.  AI sections are only included
    when non-empty content is available.
    """
    def _section(title, body):
        body = (body or "").strip()
        if not body or body.upper() in ("N/A", "NONE", ""):
            return ""
        return f"*{title}*\n{body}"

    parts = []

    desc_section = _section("Description:", description)
    if desc_section:
        parts.append(desc_section)

    steps_section = _section("Steps to Reproduce:", steps)
    if steps_section:
        parts.append(steps_section)

    expected_section = _section("Expected Result:", expected_output)
    if expected_section:
        parts.append(expected_section)

    failure_section = _section("Failure Reason:", failure_reason)
    if failure_section:
        parts.append(failure_section)

    # ── AI sections (only when content exists) ────────────────────────────
    ai_summary_clean = (ai_summary or "").strip()
    ai_fix_clean     = (ai_fix     or "").strip()

    if ai_summary_clean or ai_fix_clean:
        parts.append("----")   # Jira wiki horizontal rule as separator

    if ai_summary_clean:
        parts.append(f"*AI Summary:*\n{ai_summary_clean}")

    if ai_fix_clean:
        parts.append(f"*AI Suggested Fix:*\n{ai_fix_clean}")

    return "\n\n".join(parts)


def write_jira_hyperlink(sheet, merged_rows, jira_id_col_index, issue_url, issue_key):
    """
    Write a Jira issue hyperlink into the JiraID column.

    If the test case spans multiple rows the JiraID cells are merged to match.
    """
    hyperlink = f'=HYPERLINK("{issue_url}", "{issue_key}")'

    if len(merged_rows) > 1:
        col_letter = sheet.cell(row=1, column=jira_id_col_index).column_letter
        merge_range = f"{col_letter}{merged_rows[0]}:{col_letter}{merged_rows[-1]}"
        sheet.merge_cells(merge_range)

    sheet.cell(row=merged_rows[0], column=jira_id_col_index).value = hyperlink


# ==========================================
# HTML DASHBOARD PATCHER
# ==========================================

def patch_html_with_jira_ids(html_path: str, jira_mapping: dict) -> None:
    """
    Inject Jira issue keys and URLs into the embedded REPORT_DATA JSON inside
    the HTML dashboard so that each failed test card shows a clickable Jira link.

    *jira_mapping* is a dict keyed by test-case title:
        { "test_case_name": {"jira_key": "SCRUM-35", "jira_url": "https://..."} }
    """
    if not html_path or not jira_mapping:
        return

    html_file = Path(html_path)
    if not html_file.is_file():
        print(f"[Jira Bot] HTML file not found, skipping patch: {html_path}")
        return

    html = html_file.read_text(encoding="utf-8")

    # Extract the embedded JSON payload
    pattern = re.compile(
        r'(const REPORT_DATA\s*=\s*)(\{.*?\})(\s*;)',
        re.DOTALL
    )
    match = pattern.search(html)
    if not match:
        print("[Jira Bot] Could not locate REPORT_DATA in HTML — skipping patch.")
        return

    try:
        payload = json.loads(match.group(2))
    except json.JSONDecodeError as exc:
        print(f"[Jira Bot] Failed to parse REPORT_DATA JSON: {exc}")
        return

    updated = False
    for result in payload.get("results", []):
        test_name = result.get("test_name", "")
        if test_name in jira_mapping:
            entry = jira_mapping[test_name]
            result["jira_id"]  = entry["jira_key"]
            result["jira_url"] = entry["jira_url"]
            updated = True

    if not updated:
        print("[Jira Bot] No matching test names found in HTML — skipping patch.")
        return

    new_json = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    new_html = html[: match.start(2)] + new_json + html[match.end(2):]

    html_file.write_text(new_html, encoding="utf-8")
    print(f"[Jira Bot] HTML dashboard patched with {len(jira_mapping)} Jira link(s) → {html_path}")


# ==========================================
# MAIN EXECUTION
# ==========================================

def main():
    if len(sys.argv) < 5:
        print("Usage: python 6_generate_jira_bot.py <PROJECT_KEY> <BOARD_NAME> <SPRINT_NAME> <EXCEL_FILE> [HTML_REPORT]")
        sys.exit(1)

    jira_project_key = sys.argv[1]
    board_name       = sys.argv[2]
    sprint_name      = sys.argv[3]
    excel_file       = sys.argv[4]
    html_report      = sys.argv[5] if len(sys.argv) >= 6 else None

    # 1. Connect to Jira
    print("Connecting to Jira...")
    try:
        jira = JIRA(options={"server": config.JIRA_SERVER}, basic_auth=(config.JIRA_USERNAME, config.JIRA_PASSWORD))
    except Exception as e:
        print(f"Failed to connect to Jira: {e}")
        sys.exit(1)

    # 2. Load the Excel workbook
    print(f"Loading Excel file: {excel_file}...")
    try:
        workbook = load_workbook(excel_file)
    except Exception as e:
        print(f"Failed to load Excel file: {e}")
        sys.exit(1)

    if config.SHEET_NAME not in workbook.sheetnames:
        print(f"Skipping JIRA Bug Creation: Sheet '{config.SHEET_NAME}' is missing (Excel mismatch).")
        sys.exit(0)

    sheet = workbook[config.SHEET_NAME]

    # 3. Locate required columns
    header_row, result_col     = find_column_by_name(sheet, config.COLUMN_STATUS)
    _, title_col               = find_column_by_name(sheet, config.COLUMN_TITLE)
    _, steps_col               = find_column_by_name(sheet, config.COLUMN_STEPS)
    _, expected_col            = find_column_by_name(sheet, config.COLUMN_EXPECTED_RESULT)
    _, actual_col              = find_column_by_name(sheet, config.COLUMN_DESCRIPTION)
    _, failure_reason_col      = find_column_by_name(sheet, config.COLUMN_FAILURE_REASON)

    # AI columns — optional (warn but do not abort if missing)
    _, ai_summary_col = find_column_by_name(sheet, config.COLUMN_AI_SUMMARY)
    _, ai_fix_col     = find_column_by_name(sheet, config.COLUMN_AI_FIX)
    # Test Class column — used as fallback for HTML test_name matching
    _, test_class_col = find_column_by_name(sheet, "Test Class")

    if ai_summary_col is None:
        print(f"[Jira Bot] Note: Column '{config.COLUMN_AI_SUMMARY}' not found — AI summary will be omitted.")
    if ai_fix_col is None:
        print(f"[Jira Bot] Note: Column '{config.COLUMN_AI_FIX}' not found — AI suggested fix will be omitted.")

    # Validate required columns
    missing_cols = []
    if not header_row or not result_col:
        missing_cols.append(config.COLUMN_STATUS)
    if not title_col:
        missing_cols.append(config.COLUMN_TITLE)
    if not steps_col:
        missing_cols.append(config.COLUMN_STEPS)
    if not expected_col:
        missing_cols.append(config.COLUMN_EXPECTED_RESULT)
    if not actual_col:
        missing_cols.append(config.COLUMN_DESCRIPTION)
    if not failure_reason_col:
        missing_cols.append(config.COLUMN_FAILURE_REASON)

    if missing_cols:
        print(f"Skipping JIRA Bug Creation: Required columns {missing_cols} are missing from the sheet (Column name mismatch).")
        sys.exit(0)

    # 4. Ensure the JiraID and CreateJira columns exist
    jira_id_col = ensure_jira_id_column(sheet, header_row)
    create_jira_col = ensure_create_jira_column(sheet, header_row)

    # 5. Resolve the active sprint
    print(f"Fetching board '{board_name}' and sprint '{sprint_name}'...")
    boards = jira.boards(name=board_name)
    if not boards:
        print(f"No board found with name '{board_name}'.")
        sys.exit(1)

    sprints = jira.sprints(boards[0].id)
    active_sprint = next(
        (s for s in sprints if s.state == "active" and s.name == sprint_name), None
    )
    if not active_sprint:
        print(f"No active sprint found with name '{sprint_name}'.")
        sys.exit(1)

    # 6. Process failed test cases
    print("Scanning for failed test cases...")
    row = header_row + 1

    # Maps test_case_title → {jira_key, jira_url} for HTML patching
    jira_mapping: dict = {}

    while row <= sheet.max_row:
        result    = sheet.cell(row=row, column=result_col).value
        jira_cell = sheet.cell(row=row, column=jira_id_col)

        if str(result).strip().upper() == "FAILED" and jira_cell.value is None:
            testcase_title = get_effective_cell_value(sheet, row, title_col)

            # Check if this test case is tagged to skip JIRA creation
            create_jira_val = get_effective_cell_value(sheet, row, create_jira_col)
            if create_jira_val is not None and str(create_jira_val).strip().lower() in ("false", "no", "0", "n"):
                # Collect merged rows to skip the whole test case correctly
                (merged_rows, _, _, _, _, _, _) = collect_merged_rows(
                    sheet, row, title_col, steps_col, expected_col,
                    actual_col, failure_reason_col,
                    ai_summary_col, ai_fix_col,
                    testcase_title,
                )
                print(f"  ℹ  Skipping Jira creation for row {row} ('{testcase_title}') as {config.COLUMN_CREATE_JIRA} is marked '{create_jira_val}'.")
                row = merged_rows[-1] + 1
                continue


            # Collect merged rows first so we can advance `row` correctly in all paths
            (merged_rows, steps, expected_output, description,
             failure_reason, ai_summary, ai_fix) = collect_merged_rows(
                sheet, row, title_col, steps_col, expected_col,
                actual_col, failure_reason_col,
                ai_summary_col, ai_fix_col,
                testcase_title,
            )

            # When Test Case Name is "N/A" (no structured docstring), fall back
            # to the raw pytest function name from the Test Class column so the
            # Jira summary is still meaningful (e.g. "test_python_click").
            title_str = str(testcase_title or "").strip()
            if title_str.upper() == "N/A" and test_class_col:
                raw_class = get_effective_cell_value(sheet, merged_rows[0], test_class_col) or ""
                func_name = raw_class.split(".")[-1].strip()
                if func_name:
                    print(f"  ℹ  Row {row}: Test Case Name is 'N/A' — using function name '{func_name}' as Jira summary.")
                    title_str = func_name

            description_text = build_jira_description(
                description, steps, expected_output,
                failure_reason, ai_summary, ai_fix
            )

            issue = jira.create_issue(fields={
                "project":     {"key": jira_project_key},
                "summary":     title_str,
                "description": description_text,
                "issuetype":   {"name": config.ISSUE_TYPE},
            })
            jira.assign_issue(issue.key, config.ASSIGNEE_EMAIL)
            jira.add_issues_to_sprint(active_sprint.id, [issue.key])

            issue_url = f"{config.JIRA_SERVER}/browse/{issue.key}"
            write_jira_hyperlink(sheet, merged_rows, jira_id_col, issue_url, issue.key)

            # Build mapping under BOTH the Test Case Name AND the raw function
            # name from Test Class. HTML test_name is always the pytest function
            # name; this dual-key ensures a match in both cases.
            jira_entry = {"jira_key": issue.key, "jira_url": issue_url}
            jira_mapping[title_str] = jira_entry
            if test_class_col:
                raw_class = get_effective_cell_value(sheet, merged_rows[0], test_class_col) or ""
                func_name = raw_class.split(".")[-1].strip()
                if func_name and func_name != title_str:
                    jira_mapping[func_name] = jira_entry

            print(f"  ✔  Created {issue.key} for: {title_str}  (rows {merged_rows})")
            row = merged_rows[-1] + 1
            continue

        row += 1

    # 7. Persist Excel changes
    workbook.save(excel_file)
    print(f"\nExcel updated with Jira IDs → {excel_file}")

    # 8. Patch HTML dashboard (if path supplied and issues were created)
    if html_report and jira_mapping:
        patch_html_with_jira_ids(html_report, jira_mapping)
    elif html_report and not jira_mapping:
        print("[Jira Bot] No new issues created — HTML dashboard unchanged.")

    print("\n[Jira Bot] Done.")


if __name__ == "__main__":
    main()