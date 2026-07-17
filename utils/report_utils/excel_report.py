import re as _re
from openpyxl.utils import get_column_letter
import re
import logging
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pathlib
import os
from pathlib import Path

class MockStationData:
    def __init__(self):
        self.log_folder_path = "logs"
        self.reports_folder_path = "reports"
        self.logger = None

sd = MockStationData()
_logger = logging.getLogger(__name__)
_tc_row_map = {}


_ANSI_RE_XL = re.compile(r'\x1b\[[0-9;]*[mGKHFJA-Za-z]|\x1b\(B')

def clean_failure_reason(msg):
    """Strip ANSI codes, traceback noise and pytest diff artifacts.
    Returns a short, human-readable failure summary."""
    if not msg:
        return "Unknown Failure"

    # 1. Remove ANSI / terminal colour escape sequences
    msg = _ANSI_RE_XL.sub('', msg)

    raw_lines = msg.splitlines()
    lines = []
    for line in raw_lines:
        stripped = line.strip()
        if stripped.startswith('---') and any(
            kw in line for kw in ('Captured', 'stdout', 'stderr', 'teardown', 'log')
        ):
            break
        if any(kw in line for kw in (
            'failed with error:', 'case iddd', 'Error getting test case details'
        )):
            continue
        lines.append(line)

    # 2. Extract meaningful lines only
    assert_line = ""
    error_lines = []

    for line in lines:
        trimmed = line.strip()
        if trimmed.startswith('? '):
            continue
        if trimmed.startswith('> ') or line.startswith('>'):
            code = trimmed.lstrip('> ').strip()
            if code and not assert_line:
                assert_line = code
            continue
        if trimmed.startswith('E '):
            detail = trimmed[2:].strip()
            if detail:
                error_lines.append(detail)
            continue

    kept_error = []
    diff_lines = []
    for el in error_lines:
        if el.startswith('assert ') or el.startswith('AssertionError') or (':' in el and any(
            kw in el for kw in ('Error', 'Exception', 'Failure', 'assert')
        )):
            kept_error.append(el)
        elif el.startswith('+ ') or el.startswith('- '):
            diff_lines.append(el)
        elif not kept_error:
            kept_error.append(el)

    parts = []
    if assert_line:
        parts.append(f"Failed at: {assert_line}")
    parts.extend(kept_error[:2])
    parts.extend(diff_lines[:2])

    result = '\n'.join(parts).strip()
    if result:
        return result

    non_empty = [l.strip() for l in lines if l.strip()]
    for line in reversed(non_empty):
        if any(kw in line for kw in ('Error', 'Exception', 'assert', 'Fail')):
            return line
    return non_empty[-1] if non_empty else "Unknown Failure"

# ── SHARED EXCEL STYLES ────────────────────────────────────────────────────

_FONT_NAME = "Arial"

_THIN      = Side(style="thin")
_BDR       = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_LEFT_BDR  = Border(left=_THIN,              top=_THIN, bottom=_THIN)
_RIGHT_BDR = Border(             right=_THIN, top=_THIN, bottom=_THIN)
_INNER_BDR = Border(                          top=_THIN, bottom=_THIN)

_ALIGN_CTR     = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
_ALIGN_LEFT    = Alignment(horizontal="left",   vertical="bottom", wrap_text=True)
_ALIGN_TOP     = Alignment(horizontal="left",   vertical="bottom", wrap_text=True)
_ALIGN_CTR_MID = Alignment(horizontal="center", vertical="center", wrap_text=True)

_FILL_PASS    = PatternFill("solid", fgColor="C6EFCE")
_FILL_FAIL    = PatternFill("solid", fgColor="FFC7CE")
_FILL_EVEN    = PatternFill("solid", fgColor="EBF3FB")
_FILL_ODD     = PatternFill("solid", fgColor="FFFFFF")
_FILL_TITLE   = PatternFill("solid", fgColor="1F3864")
_FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
_FILL_SUBHDR  = PatternFill("solid", fgColor="D6E4F0")
_FILL_LABEL   = PatternFill("solid", fgColor="D6E4F0")
_FILL_EVEN2   = PatternFill("solid", fgColor="EBF3FB")
_FILL_SKIP    = PatternFill("solid", fgColor="F2E6FF")
_FILL_NODATA  = PatternFill("solid", fgColor="FFF2CC")

_FONT_TITLE   = Font(name=_FONT_NAME, bold=True,  size=14, color="FFFFFF")
_FONT_SECTION = Font(name=_FONT_NAME, bold=True,  size=11, color="FFFFFF")
_FONT_SUBHDR  = Font(name=_FONT_NAME, bold=True,  size=10, color="000000")
_FONT_LABEL   = Font(name=_FONT_NAME, bold=True,  size=10)
_FONT_LINK    = Font(name=_FONT_NAME,              size=10, color="0563C1", underline="single")
_FONT_ERROR   = Font(name=_FONT_NAME, bold=True,   size=10, color="FF0000")
_FONT_SKIP    = Font(name=_FONT_NAME, bold=True,   size=10, color="7030A0")
_FONT_PASS    = Font(name=_FONT_NAME, bold=True,   size=10, color="375623")
_FONT_FAIL    = Font(name=_FONT_NAME, bold=True,   size=10, color="9C0006")
_FONT_NODATA  = Font(name=_FONT_NAME, italic=True, size=10, color="7F6000", bold=True)
_FONT_DATA    = Font(name=_FONT_NAME, size=10)

# ── Helpers ────────────────────────────────────────────────────────────────

def _safe_str(value) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    text = _re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    return text

def _no_data_row(ws, row: int, col_start: int, col_end: int, message: str) -> None:
    _merge_row(ws, row, col_start, col_end, message, _FONT_NODATA, _FILL_NODATA)
    ws.row_dimensions[row].height = 22

def _auto_row_height(values: list, col_widths: list, min_h: int = 20) -> float:
    max_lines = 1
    for val, width in zip(values, col_widths):
        if not val:
            continue
        chars = max(int(width * 1.1), 1)
        lines = sum(
            max(1, (len(p) + chars - 1) // chars)
            for p in str(val).split("\n")
        )
        max_lines = max(max_lines, lines)
    return max(min_h, min(max_lines * 13 + 4, 200))

def _hdr_cell(ws, row, col, value, fill_hex="1F3864"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
    c.fill      = PatternFill("solid", fgColor=fill_hex)
    c.alignment = _ALIGN_CTR_MID
    c.border    = _BDR
    return c

def _data_cell(ws, row, col, value, align=None, number_format=None):
    safe_val = _safe_str(value) if isinstance(value, str) else value
    c = ws.cell(row=row, column=col, value=safe_val)
    c.font      = _FONT_DATA
    c.alignment = align if align is not None else _ALIGN_LEFT
    c.border    = _BDR
    if number_format:
        c.number_format = number_format
    return c

def _merge_row(ws, row, col_start, col_end, value, font, fill, align=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    span = col_end - col_start + 1
    for ci in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=ci)
        c.fill      = fill
        c.font      = font
        c.alignment = align or _ALIGN_CTR_MID
        if span == 1:
            c.border = _BDR
        elif ci == col_start:
            c.border = _LEFT_BDR
        elif ci == col_end:
            c.border = _RIGHT_BDR
        else:
            c.border = _INNER_BDR
    ws.cell(row=row, column=col_start).value = value

def _s2_cell(ws, row, col, value, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font  or _FONT_DATA
    c.fill      = fill  or _FILL_ODD
    c.alignment = align or _ALIGN_LEFT
    c.border    = _BDR
    return c

def _screenshot_cell(ws, row, col, screenshot_path: str, row_fill):
    sc = ws.cell(row=row, column=col)
    sc.fill   = row_fill
    sc.border = _BDR
    if screenshot_path:
        rel = str(
            pathlib.Path("screenshots") / pathlib.Path(screenshot_path).name
        ).replace("\\", "/")
        sc.value     = "View Screenshot"
        sc.hyperlink = rel
        sc.font      = Font(name=_FONT_NAME, color="003366",
                            underline="single", bold=True, size=10)
        sc.alignment = _ALIGN_CTR
    else:
        sc.value     = "—"
        sc.font      = Font(name=_FONT_NAME, size=10, color="999999")
        sc.alignment = _ALIGN_CTR

# ── UI DETAILS SHEET ───────────────────────────────────────────────────────

_UI_COL_WIDTHS = [
    12,   # Platform
    30,   # Feature
    55,   # Test Case Name
    20,   # Test Case ID
    20,   # Module
    45,   # Description
    45,   # Steps
    45,   # Expected Output
    14,   # Suite
    40,   # Test Class
    12,   # Status
    14,   # Duration
    60,   # Failure Reason
    60,   # Suggested Fix
    60,   # Short Summary
    25,   # Screenshot
    26    # Executed At
]

_UI_HEADERS = [
    "Platform",
    "Feature",
    "Test Case Name",
    "Test Case ID",
    "Module",
    "Description",
    "Steps",
    "Expected Output",
    "Suite",
    "Test Class",
    "Status",
    "Duration (s)",
    "Failure Reason",
    "Suggested Fix",
    "Short Summary",
    "Screenshot",
    "Executed At",
]

_UI_STATUS_COL   = 11
_UI_MESSAGE_COL  = 13
_UI_FIX_COL      = 14
_UI_SUMMARY_COL  = 15
_UI_SCSHOT_COL   = 16
def _format_duration(total_seconds: float) -> str:
    total_seconds = round(total_seconds, 1)
    if total_seconds <= 0:
        return "—"
    if total_seconds < 60:
        return f"{total_seconds}s"
    total_int = int(total_seconds)
    m, s = divmod(total_int, 60)
    if m < 60:
        return f"{m}m {s}s"
    h, m = divmod(m, 60)
    return f"{h}h {m}m {s}s"

def _write_ui_details_sheet(ws, results: list) -> None:
    for i, w in enumerate(_UI_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, hdr in enumerate(_UI_HEADERS, start=1):
        _hdr_cell(ws, 1, col, hdr, fill_hex="1F3864")
    ws.row_dimensions[1].height = 22

    if not results:
        _no_data_row(ws, 2, 1, len(_UI_HEADERS),
                     "⚠  No test cases were executed for this module.")
        ws.freeze_panes = "A2"
        return

    for row_idx, r in enumerate(results, start=2):
        status = (r.get("status") or "").upper()

        if r.get("screenshot") and status == "PASSED":
            status      = "FAILED"
            r["status"] = "FAILED"
            if not r.get("message"):
                r["message"] = "Test captured a failure screenshot — status corrected to FAILED."

        if status == "PASSED":
            row_fill = _FILL_PASS
        elif status == "FAILED":
            row_fill = _FILL_FAIL
        elif status == "SKIPPED":
            row_fill = _FILL_SKIP
        else:
            row_fill = _FILL_EVEN if row_idx % 2 == 0 else _FILL_ODD

        failure_reason = _safe_str(r.get("message", "")) if status == "FAILED" else ""

        values = [
            r.get("platform",    "Web"),
            r.get("feature",     ""),
            r.get("test_name",   ""),
            r.get("tc_id",       ""),
            r.get("module",      ""),
            r.get("description", ""),
            r.get("steps",       ""),
            r.get("expected",    ""),
            r.get("suite",       "pytest"),
            r.get("test_class",  ""),
            status,
            r.get("duration",    ""),
            failure_reason,
            r.get("Suggested Fix", ""),
            r.get("Short Summary", ""),
            "",
            r.get("executed_at", ""),
        ]

        for col, val in enumerate(values, start=1):
            align = _ALIGN_CTR if col in (1, 2, 4, 9, 11, 12, 15) else _ALIGN_LEFT
            c     = _data_cell(ws, row_idx, col, val, align=align)
            c.fill = row_fill

            if col == _UI_STATUS_COL:
                if status == "PASSED":
                    c.font = _FONT_PASS
                elif status == "FAILED":
                    c.font = _FONT_FAIL
                elif status == "SKIPPED":
                    c.font = _FONT_SKIP
                else:
                    c.font = Font(name=_FONT_NAME, bold=True, size=10)

            if col == _UI_MESSAGE_COL and status == "FAILED" and val:
                if col in (_UI_FIX_COL, _UI_SUMMARY_COL) and val:

                    c.alignment = _ALIGN_LEFT

                    if status == "FAILED":
                    
                        c.font = Font(
                            name=_FONT_NAME,
                            size=10,
                            color="9C0006"
                        )
                c.font = Font(name=_FONT_NAME, bold=True, size=10, color="9C0006")

            if col == 12 and val != "":
                c.number_format = "0.00"

        _screenshot_cell(ws, row_idx, _UI_SCSHOT_COL, r.get("screenshot", ""), row_fill)
        ws.row_dimensions[row_idx].height = _auto_row_height(values, _UI_COL_WIDTHS)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(_UI_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(results) + 1}"

# ── UI SUMMARY SHEET ───────────────────────────────────────────────────────

def _write_ui_summary_sheet(ws, module_results: dict,
                             exec_seconds: float = 0.0) -> None:
    for col, w in zip(["A","B","C","D","E","F","G"],
                      [30,  14,  14,  10,  10,  10,  14]):
        ws.column_dimensions[col].width = w

    row = 1
    _merge_row(ws, row, 1, 7, "UI Test Summary Report", _FONT_TITLE, _FILL_TITLE)
    ws.row_dimensions[row].height = 40
    row += 1

    _all_results_flat = [r for v in module_results.values() for r in v]
    _total_secs = sum(r.get("duration", 0) or 0 for r in _all_results_flat)

    # ── FIX: exec time only shown when UI tests actually ran ──────────────
    exec_time_str = (
        _format_duration(exec_seconds) if exec_seconds > 0
        else _format_duration(_total_secs) if _total_secs > 0
        else "—"
    )

    for label, value in [
        ("Date",                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Suite",                "pytest"),
        ("Total Execution Time", exec_time_str),
    ]:
        _s2_cell(ws, row, 1, label, font=_FONT_LABEL, fill=_FILL_LABEL,
                 align=_ALIGN_CTR_MID)
        _s2_cell(ws, row, 2, value,
                 font=Font(name=_FONT_NAME, size=10, bold=True),
                 fill=_FILL_ODD, align=_ALIGN_CTR_MID)
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

    # Feature summary table
    _merge_row(ws, row, 1, 7, "Feature Summary", _FONT_SECTION, _FILL_SECTION)
    ws.row_dimensions[row].height = 26
    row += 1

    for col, hdr in enumerate(
        ["Feature", "FAILED", "PASSED", "ERROR", "SKIPPED", "Total", "Pass Rate"],
        start=1
    ):
        c = _s2_cell(ws, row, col, hdr, font=_FONT_SUBHDR,
                     fill=_FILL_SUBHDR, align=_ALIGN_CTR_MID)
    ws.row_dimensions[row].height = 22
    row += 1

    grand = {"failed": 0, "passed": 0, "error": 0, "skipped": 0}

    for mi, (feature, results) in enumerate(module_results.items()):
        rfill = _FILL_EVEN2 if mi % 2 == 0 else _FILL_ODD

        if not results:
            _merge_row(ws, row, 1, 7,
                       f"⚠  {feature}  —  No test cases were executed.",
                       _FONT_NODATA, _FILL_NODATA)
            ws.row_dimensions[row].height = 22
            row += 1
            continue

        for r in results:
            if r.get("screenshot") and (r.get("status") or "").upper() == "PASSED":
                r["status"] = "FAILED"
                if not r.get("message"):
                    r["message"] = "Test captured a failure screenshot — status corrected to FAILED."

        failed  = sum(1 for r in results if (r.get("status") or "").upper() == "FAILED")
        passed  = sum(1 for r in results if (r.get("status") or "").upper() == "PASSED")
        error   = sum(1 for r in results if (r.get("status") or "").upper() == "ERROR")
        skipped = sum(1 for r in results if (r.get("status") or "").upper() == "SKIPPED")
        total   = failed + passed + error + skipped
        pct     = (passed / total * 100) if total else 0.0

        grand["failed"]  += failed
        grand["passed"]  += passed
        grand["error"]   += error
        grand["skipped"] += skipped

        pct_fill = (_FILL_PASS if pct == 100
                    else _FILL_FAIL if pct == 0 and total > 0
                    else rfill)

        for col, val in enumerate(
            [feature, failed, passed, error, skipped, total, f"{pct:.2f}%"], start=1
        ):
            _s2_cell(ws, row, col, val,
                     font=Font(name=_FONT_NAME, size=10, bold=(col == 1)),
                     fill=pct_fill if col == 7 else rfill,
                     align=_ALIGN_CTR)
        ws.row_dimensions[row].height = 22
        row += 1

    g_total = sum(grand.values())
    g_pct   = (grand["passed"] / g_total * 100) if g_total else 0.0
    g_pfill = (_FILL_PASS if g_pct == 100
               else _FILL_FAIL if g_pct == 0 and g_total > 0
               else _FILL_EVEN2)

    for col, val in enumerate(
        ["TOTAL", grand["failed"], grand["passed"], grand["error"],
         grand["skipped"], g_total, f"{g_pct:.2f}%"],
        start=1,
    ):
        _s2_cell(ws, row, col, val,
                 font=Font(name=_FONT_NAME, size=10, bold=True),
                 fill=g_pfill if col == 7 else _FILL_SUBHDR,
                 align=_ALIGN_CTR)
    ws.row_dimensions[row].height = 22
    row += 2

    # Failed test cases
    _merge_row(ws, row, 1, 6, "Failed Test Cases", _FONT_SECTION, _FILL_SECTION)
    ws.row_dimensions[row].height = 26
    row += 1

    for col, hdr in enumerate(
        ["Feature", "Test Case Name", "TC ID", "Module", "Failure Reason", "Screenshot"],
        start=1,
    ):
        _s2_cell(ws, row, col, hdr, font=_FONT_SUBHDR,
                 fill=_FILL_SUBHDR, align=_ALIGN_CTR_MID)
    ws.row_dimensions[row].height = 22
    row += 1

    all_ui = [r for v in module_results.values() for r in v]
    if not all_ui:
        _no_data_row(ws, row, 1, 6,
                     "⚠  No test cases were executed — nothing to report.")
        row += 1
    else:
        any_failed = False
        for feature, results in module_results.items():
            for r in (x for x in results if (x.get("status") or "").upper() == "FAILED"):
                any_failed = True
                failure_reason = _safe_str(r.get("message", "")) or "—"
                for col, val in enumerate(
                    [feature, r.get("test_name",""), r.get("tc_id",""),
                     r.get("module",""), failure_reason, ""],
                    start=1,
                ):
                    align = _ALIGN_LEFT if col == 5 else _ALIGN_CTR
                    _s2_cell(
                        ws, row, col, val,
                        font=Font(name=_FONT_NAME, size=10,
                                  bold=(col == 1),
                                  color="9C0006" if col == 5 else "000000"),
                        fill=_FILL_FAIL, align=align,
                    )
                _screenshot_cell(ws, row, 6, r.get("screenshot",""), _FILL_FAIL)
                ws.row_dimensions[row].height = max(30, min(len(failure_reason)//3, 80))
                row += 1

        if not any_failed:
            _merge_row(ws, row, 1, 6, "✔  No failed test cases.",
                       Font(name=_FONT_NAME, italic=True, size=10, color="375623"),
                       _FILL_PASS)
            ws.row_dimensions[row].height = 22

    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 20

# ── CHATBOT DETAILS SHEET ──────────────────────────────────────────────────

_CB_COL_WIDTHS = [
    10, 45, 45, 45, 16,
    14, 50, 14, 50,
    14, 50, 16, 55,
    20, 22,
]
_CB_HEADERS = [
    "ID", "Input", "Expected Output", "Actual Output", "Result",
    "GEval Score",        "GEval Reason",
    "Faithfulness Score", "Faithfulness Reason",
    "Relevancy Score",    "Relevancy Reason",
    "Completeness Score", "Completeness Reason",
    "Final Score",        "Screenshot",
]
_CB_SCREENSHOT_COL = 15

def _write_results_sheet(ws, results: list, screenshots: dict) -> None:
    for i, width in enumerate(_CB_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for col, header in enumerate(_CB_HEADERS, start=1):
        _hdr_cell(ws, 1, col, header)
    ws.row_dimensions[1].height = 22

    if not results:
        _no_data_row(ws, 2, 1, len(_CB_HEADERS),
                     "⚠  No test cases were executed for this module.")
        ws.freeze_panes = "A2"
        return

    for row_idx, r in enumerate(results, start=2):
        _tc_row_map[r.get("id")] = row_idx
        is_error   = bool(r.get("error"))
        is_skipped = r.get("result") == "SKIPPED"

        row_fill = (
            _FILL_FAIL if r["result"] == "FAIL"
            else (_FILL_EVEN if row_idx % 2 == 0 else _FILL_ODD)
        )

        row_values = [
            r.get("id"),               r.get("input"),
            r.get("expected_output"),  r.get("actual_output"),
            r.get("result"),
            r.get("geval_score"),          r.get("geval_reason"),
            r.get("faithfulness_score"),   r.get("faithfulness_reason"),
            r.get("relevancy_score"),      r.get("relevancy_reason"),
            r.get("completeness_score"),   r.get("completeness_reason"),
            r.get("final_score"),
        ]

        metric_cells = []

        for col, val in enumerate(row_values[:5], start=1):
            c = _data_cell(ws, row_idx, col, val,
                           align=_ALIGN_CTR if col in (1, 5) else _ALIGN_LEFT)
            c.fill = row_fill
            if col == 5:
                if val == "PASS":
                    c.fill = _FILL_PASS
                    c.font = _FONT_PASS
                elif val == "FAIL":
                    c.fill = _FILL_FAIL
                    c.font = _FONT_FAIL

        for col in range(6, 14):
            val   = row_values[col - 1]
            align = _ALIGN_CTR if col in (6, 8, 10, 12) else _ALIGN_LEFT
            fmt   = "0.00"     if col in (6, 8, 10, 12) else None
            c = _data_cell(ws, row_idx, col, val, align=align, number_format=fmt)
            c.fill = row_fill
            metric_cells.append(c)

        c14 = _data_cell(ws, row_idx, 14, r.get("final_score"),
                         align=_ALIGN_CTR, number_format="0.0000")
        c14.fill = row_fill
        metric_cells.append(c14)

        if is_error:
            for c in metric_cells:
                c.font = _FONT_ERROR
        elif is_skipped:
            for c in metric_cells:
                c.font = _FONT_SKIP
                if c.value in (None, ""):
                    c.value = "SKIPPED (weight=0)"

        tc_id    = r.get("id", "")
        filename = screenshots.get(tc_id, "")
        _screenshot_cell(ws, row_idx, _CB_SCREENSHOT_COL, filename,
                         PatternFill("solid", fgColor="FFFFFF"))

        visible = row_values + [filename]
        ws.row_dimensions[row_idx].height = _auto_row_height(visible, _CB_COL_WIDTHS)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(_CB_HEADERS))}{len(results) + 1}"
    )

# ── CHATBOT SUMMARY SHEET ──────────────────────────────────────────────────

def _write_summary_sheet(ws2, module_results: dict, threshold: float,
                          exec_seconds: float = 0.0, category_weights: dict = None, judge_model: str = "") -> None:
    for col, width in zip(
        ["A","B","C","D","E","F","G","H","I","J"],
        [22,  22,  22,  18,  18,  18,  12,  14,  14,  14],
    ):
        ws2.column_dimensions[col].width = width

    all_results = [r for lst in module_results.values() for r in lst]
    row = 1

    _merge_row(ws2, row, 1, 5, "Chatbot Test Summary Report",
               _FONT_TITLE, _FILL_TITLE)
    ws2.row_dimensions[row].height = 40
    row += 1

    # ── FIX: exec time only shown when chatbot tests actually ran ─────────
    exec_time_str = (
        _format_duration(exec_seconds) if (exec_seconds > 0 and all_results)
        else "—"
    )

    for label, value in [
        ("Version",               "Not Applicable"),
        ("Release",               "Not Applicable"),
        ("Date",                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Evaluation Model",      judge_model),
        ("Total Execution Time",  exec_time_str),
    ]:
        _s2_cell(ws2, row, 1, label,  font=_FONT_LABEL, fill=_FILL_LABEL,
                 align=_ALIGN_CTR_MID)
        _s2_cell(ws2, row, 2, value,
                 font=Font(name=_FONT_NAME, size=10, bold=True),
                 fill=_FILL_ODD, align=_ALIGN_CTR_MID)
        ws2.row_dimensions[row].height = 22
        row += 1

    row += 1

    # Module summary table
    _merge_row(ws2, row, 1, 5, "Module Summary", _FONT_SECTION, _FILL_SECTION)
    ws2.row_dimensions[row].height = 26
    row += 1

    for col, hdr in enumerate(
        ["Module", "Total Tests", "Passed", "Failed", "Pass Rate"], start=1
    ):
        _s2_cell(ws2, row, col, hdr, font=_FONT_SUBHDR,
                 fill=_FILL_SUBHDR, align=_ALIGN_CTR_MID)
    ws2.row_dimensions[row].height = 22
    row += 1

    for mi, (sheet_name, results) in enumerate(module_results.items()):
        total_m = len(results)
        rfill   = _FILL_EVEN2 if mi % 2 == 0 else _FILL_ODD

        if total_m == 0:
            _merge_row(ws2, row, 1, 5,
                       f"⚠  {sheet_name}  —  No test cases were executed.",
                       _FONT_NODATA, _FILL_NODATA)
            ws2.row_dimensions[row].height = 22
            row += 1
            continue

        passed_m = sum(1 for r in results if r["result"] == "PASS")
        failed_m = total_m - passed_m
        pct_m    = (passed_m / total_m * 100) if total_m else 0
        pct_fill = (_FILL_PASS if pct_m == 100
                    else _FILL_FAIL if pct_m == 0 and total_m > 0
                    else rfill)

        for col, val in enumerate(
            [sheet_name, total_m, passed_m, failed_m, f"{pct_m:.2f}%"], start=1
        ):
            _s2_cell(ws2, row, col, val,
                     font=Font(name=_FONT_NAME, size=10, bold=(col == 1)),
                     fill=pct_fill if col == 5 else rfill,
                     align=_ALIGN_CTR)
        ws2.row_dimensions[row].height = 22
        row += 1

    _s2_cell(ws2, row, 1, "Threshold Used", font=_FONT_LABEL,
             fill=_FILL_LABEL, align=_ALIGN_CTR_MID)
    _s2_cell(ws2, row, 2, threshold,
             font=Font(name=_FONT_NAME, size=10),
             fill=_FILL_ODD, align=_ALIGN_CTR_MID)
    ws2.row_dimensions[row].height = 22
    row += 2

    # Failed test cases
    _merge_row(ws2, row, 1, 3, "Failed Test Cases", _FONT_SECTION, _FILL_SECTION)
    ws2.row_dimensions[row].height = 26
    row += 1

    for col, hdr in enumerate(["Module", "TC ID", "TC Input"], start=1):
        _s2_cell(ws2, row, col, hdr, font=_FONT_SUBHDR,
                 fill=_FILL_SUBHDR, align=_ALIGN_CTR_MID)
    ws2.row_dimensions[row].height = 22
    row += 1

    if not all_results:
        _no_data_row(ws2, row, 1, 3,
                     "⚠  No test cases were executed — nothing to report.")
        row += 1
    else:
        failed_any = False
        gfi = 0
        for sheet_name, results in module_results.items():
            for r in (x for x in results if x["result"] == "FAIL"):
                failed_any = True
                rfill = _FILL_EVEN2 if gfi % 2 == 0 else _FILL_ODD
                _s2_cell(ws2, row, 1, sheet_name,
                         font=_FONT_DATA, fill=rfill, align=_ALIGN_CTR)
                tc_id   = r["id"]
                lc      = ws2.cell(row=row, column=2, value=tc_id)
                row_num = _tc_row_map.get(tc_id, 1)
                lc.hyperlink  = f"#{sheet_name}_Chatbot!A{row_num}"
                lc.font       = _FONT_LINK
                lc.fill       = rfill
                lc.alignment  = _ALIGN_CTR
                lc.border     = _BDR
                _s2_cell(ws2, row, 3, r.get("input",""),
                         font=_FONT_DATA, fill=rfill, align=_ALIGN_CTR)
                ws2.row_dimensions[row].height = 22
                row += 1
                gfi += 1

        if not failed_any:
            _merge_row(ws2, row, 1, 3, "✔  No failed test cases.",
                       Font(name=_FONT_NAME, italic=True, size=10, color="375623"),
                       _FILL_PASS)
            ws2.row_dimensions[row].height = 22
            row += 1

    row += 1

    _merge_row(ws2, row, 1, 4, "Metrics Weights", _FONT_SECTION, _FILL_SECTION)
    ws2.row_dimensions[row].height = 26
    row += 1

    for col, hdr in enumerate(
        ["GEval", "Faithfulness", "Relevancy", "Completeness"],
        start=1,
    ):
        _s2_cell(ws2, row, col, hdr, font=_FONT_SUBHDR,
                 fill=_FILL_SUBHDR, align=_ALIGN_CTR_MID)
    ws2.row_dimensions[row].height = 22
    row += 1

    if not all_results:
        _no_data_row(ws2, row, 1, 4,
                     "⚠  No test cases were executed — no weights to display.")
        row += 1
    else:
        weights = category_weights.get("ALL", (0, 0, 0, 0))
        for col, val in enumerate(weights, start=1):
            _s2_cell(ws2, row, col, val,
                     font=_FONT_DATA,
                     fill=_FILL_EVEN2,
                     align=_ALIGN_CTR)
        ws2.row_dimensions[row].height = 22
        row += 1

    ws2.freeze_panes = None


def write_excel_report(judge_model, category_weights, screenshots_dict: dict, module_results_registry: dict, output_path: str,
                        threshold: float, exec_seconds: float = 0.0) -> None:
    screenshots = screenshots_dict
    try:
        from tests.web.reports_tab.test_reportiq_chat_with_docs import _screenshot_map
        screenshots.update({
            tc_id: pathlib.Path(path).name
            for tc_id, path in _screenshot_map.items()
            if path
        })
    except ImportError:
        pass

    chatbot_modules = {
        n: v for n, v in module_results_registry.get("CHATBOT", {}).items() if v
    }
    ui_modules = {
        n: v for n, v in module_results_registry.get("UI", {}).items() if v
    }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if chatbot_modules:
        ws_chat_summary = wb.create_sheet("Chatbot Summary")
        _write_summary_sheet(ws_chat_summary, chatbot_modules, threshold,
                             exec_seconds=exec_seconds, category_weights=category_weights, judge_model=judge_model)

    if ui_modules:
        ws_ui_summary = wb.create_sheet("UI Summary")
        _write_ui_summary_sheet(ws_ui_summary, ui_modules,
                                exec_seconds=exec_seconds)

    for sheet_name, results in chatbot_modules.items():
        safe_name      = f"{sheet_name}_Chatbot"[:31]
        ws             = wb.create_sheet(safe_name)
        sorted_results = sorted(results, key=lambda r: r.get("result") != "FAIL")
        _write_results_sheet(ws, sorted_results, screenshots)

    for sheet_name, results in ui_modules.items():
        safe_name      = f"{sheet_name}_UI"[:31]
        ws             = wb.create_sheet(safe_name)
        sorted_results = sorted(
            results,
            key=lambda r: (
                0 if (r.get("status") or "").upper() == "FAILED" else 1,
                r.get("feature", ""),
            ),
        )
        _write_ui_details_sheet(ws, sorted_results)

    wb.save(output_path)
    print(f"  Excel saved → {output_path}")


BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

ROOT_LOG_DIR = os.path.join(
    BASE_DIR,
    "logs"
)

os.makedirs(
    ROOT_LOG_DIR,
    exist_ok=True
)

def _parse_junit_xml(xml_file, root_log_dir, ai_state=None):
    import xml.etree.ElementTree as ET

    if not os.path.exists(xml_file) or os.path.getsize(xml_file) == 0:
        return None

    root = ET.parse(xml_file).getroot()
    results = []

    for suite in root.findall(".//testsuite"):
        for case in suite.findall("testcase"):
            classname = case.get("classname")
            raw_name = case.get("name")

            if classname in ("internal", None) or raw_name in ("internal", None):
                continue

            module_info = _extract_module_from_classname(classname)
            doc_info = _parse_test_docstring(classname, raw_name)

            if not doc_info.get("Test Case Name") or str(doc_info.get("Test Case Name", "")).strip() in ("N/A", "", "None"):
                doc_info["Test Case Name"] = raw_name.split("[")[0]

            if "[" in raw_name and "]" in raw_name:
                param = raw_name.split("[", 1)[1].rstrip("]")
                doc_info["Test Case Name"] += f" [{param}]"

            status = "PASSED"
            message = "Test passed successfully"

            if case.find("failure") is not None:
                status = "FAILED"
                fail_elem = case.find("failure")
                raw_msg = fail_elem.text or fail_elem.get("message", "")
                message = clean_failure_reason(raw_msg)
            elif case.find("error") is not None:
                status = "ERROR"
                err_elem = case.find("error")
                raw_msg = err_elem.text or err_elem.get("message", "")
                message = clean_failure_reason(raw_msg)
            elif case.find("skipped") is not None:
                status = "SKIPPED"
                skip_elem = case.find("skipped")
                raw_msg = skip_elem.text or skip_elem.get("message", "")
                message = clean_failure_reason(raw_msg)
            duration = round(float(case.get("time", 0)), 2)

            global WEB_EXECUTION_TIME
            if module_info["Platform"] == "Web":
                WEB_EXECUTION_TIME += duration

            screenshot_path = ""
            properties = case.find("properties")
            if properties is not None:
                for prop in properties.findall("property"):
                    if prop.get("name") == "SCREENSHOT":
                        screenshot_path = os.path.basename(prop.get("value"))
                        break

            # Defaults
            suggested_fix = ""
            short_summary = ""
            jira_id = ""
            jira_link = ""

            try:
                # =====================================================
                # GET TEST FILE NAME — handles both test types:
                #   Function-based: tests.apps.web_app.programiz.test_programiz
                #     -> last part IS the module file  -> use [-1]
                #   Class-based:    tests.apps.product.test_product.TestProductScenarios
                #     -> last part is the class name   -> use [-2]
                # =====================================================
                _parts = classname.split(".")
                _last  = _parts[-1] if _parts else ""
                if _last and _last[0].isupper():
                    # Class-based test — module file is second-to-last
                    test_module_file = _parts[-2]
                else:
                    # Function-based test — last segment IS the module file
                    test_module_file = _last

                test_case_folder = raw_name.split("[")[0].strip()
                ai_summary_file = os.path.join(
                    root_log_dir,
                    test_case_folder,
                    "ai_failure_summary.txt"
                )
                # Do not create empty directories for test cases. Only read the summary if it exists.

                if status in ("FAILED", "ERROR"):
                    if ai_state:
                        # Extract mapping lists
                        jira_map = {j["test_name"]: j for j in ai_state.get("jira_results", [])}
                        jira_healed_map_fail = {j["test_name"]: j for j in ai_state.get("jira_results_healed", [])}
                        class_map = {c["test_name"]: c for c in ai_state.get("failure_classifications", [])}
                        rec_map = {r["test_name"]: r for r in ai_state.get("action_recommendations", [])}

                        matched_test_name = None

                        def match_test_name(cname, rname, state_name):
                            if not state_name:
                                return False
                            cname_str = cname or ""
                            rname_str = rname or ""
                            if state_name.endswith('.py'):
                                state_module = state_name.replace('.py', '').replace('/', '.')
                                if state_module in cname_str or state_module in rname_str:
                                    return True
                            last_part = state_name.split("::")[-1]
                            if rname_str != last_part:
                                return False
                            class_parts = cname_str.split(".")
                            for part in class_parts:
                                if part and part[0].islower() and part not in state_name:
                                    return False
                            return True

                        for state_test_name in class_map.keys():
                            if match_test_name(classname, raw_name, state_test_name):
                                matched_test_name = state_test_name
                                break
                        if not matched_test_name:
                            for state_test_name in rec_map.keys():
                                if match_test_name(classname, raw_name, state_test_name):
                                    matched_test_name = state_test_name
                                    break
                        if not matched_test_name:
                            for state_test_name in jira_map.keys():
                                if match_test_name(classname, raw_name, state_test_name):
                                    matched_test_name = state_test_name
                                    break
                        if not matched_test_name:
                            for state_test_name in jira_healed_map_fail.keys():
                                if match_test_name(classname, raw_name, state_test_name):
                                    matched_test_name = state_test_name
                                    break

                        if matched_test_name:
                            if matched_test_name in class_map:
                                short_summary = class_map[matched_test_name].get("reasoning", "")
                            if matched_test_name in rec_map:
                                suggested_fix = rec_map[matched_test_name].get("suggested_fix", "")
                                if rec_map[matched_test_name].get("summary") and not short_summary:
                                    short_summary = rec_map[matched_test_name].get("summary")
                            if matched_test_name in jira_map:
                                jira_id = jira_map[matched_test_name].get("jira_id", "")
                                jira_link = jira_map[matched_test_name].get("jira_url", "")
                            # For healed tests that still have FAILED status in XML,
                            # fall back to jira_results_healed for Jira link
                            if (not jira_id) and matched_test_name in jira_healed_map_fail:
                                jira_id = jira_healed_map_fail[matched_test_name].get("jira_id", "")
                                jira_link = jira_healed_map_fail[matched_test_name].get("jira_url", "")

                    if not suggested_fix and not short_summary and os.path.exists(ai_summary_file):
                        with open(ai_summary_file, "r", encoding="utf-8") as f:
                            ai_text = f.read()

                        fix_match = re.search(
                            r"####\s*3\.\s*Suggested Fix\s*(.*?)(?:####\s*4\.|\Z)",
                            ai_text,
                            re.DOTALL,
                        )
                        if fix_match:
                            suggested_fix = (
                                fix_match.group(1)
                                .replace("3. Suggested Fix", "")
                                .strip()
                            )

                        summary_match = re.search(
                            r"####\s*4\.\s*Short Summary\s*(.*)",
                            ai_text,
                            re.DOTALL,
                        )
                        if summary_match:
                            short_summary = (
                                summary_match.group(1)
                                .replace("4. Short Summary", "")
                                .strip()
                            )

                elif status == "PASSED":
                    suggested_fix = "No fix required"
                    short_summary = "Test executed successfully without any failures."
                    if ai_state:
                        jira_healed_map = {j["test_name"]: j for j in ai_state.get("jira_results_healed", [])}
                        matched_test_name = None
                        def match_test_name(cname, rname, state_name):
                            if not state_name:
                                return False
                            cname_str = cname or ""
                            rname_str = rname or ""
                            if state_name.endswith('.py'):
                                state_module = state_name.replace('.py', '').replace('/', '.')
                                if state_module in cname_str or state_module in rname_str:
                                    return True
                            last_part = state_name.split("::")[-1]
                            if rname_str != last_part:
                                return False
                            class_parts = cname_str.split(".")
                            for part in class_parts:
                                if part and part[0].islower() and part not in state_name:
                                    return False
                            return True
                        for state_test_name in jira_healed_map.keys():
                            if match_test_name(classname, raw_name, state_test_name):
                                matched_test_name = state_test_name
                                break
                            # simple match fallback
                            if state_test_name.endswith('.py'):
                                state_module = state_test_name.replace('.py', '').replace('/', '.')
                                if state_module in (classname or "") or state_module in (raw_name or ""):
                                    matched_test_name = state_test_name
                                    break
                            elif (raw_name and raw_name in state_test_name) or (raw_name and state_test_name in raw_name):
                                matched_test_name = state_test_name
                                break
                        if matched_test_name:
                            jira_id = jira_healed_map[matched_test_name].get("jira_id", "")
                            jira_link = jira_healed_map[matched_test_name].get("jira_url", "")
                            suggested_fix = "Auto-healed by RegressionAI."
                            short_summary = f"Test failed initially but was successfully auto-healed. Tracked in {jira_id}."

            except Exception as e:
                print(f"AI Summary Parse Failed: {e}")

            # PR Link lookup
            pr_link_app = ""
            pr_link_tests = ""
            is_healed = False
            if ai_state:
                jira_healed_map = {j["test_name"]: j for j in ai_state.get("jira_results_healed", [])}
                for state_test_name in jira_healed_map.keys():
                    if state_test_name.endswith('.py'):
                        state_module = state_test_name.replace('.py', '').replace('/', '.')
                        if state_module in (classname or "") or state_module in (raw_name or ""):
                            is_healed = True
                            break
                    elif raw_name in state_test_name or state_test_name in raw_name:
                        is_healed = True
                        break
            if is_healed:
                if ai_state.get("pr_links"):
                    # Only store real /pull/ URLs — filter out any /tree/ placeholders
                    real_prs = [str(u) for u in ai_state.get("pr_links") if u and "/pull/" in str(u)]
                    for u in real_prs:
                        if "agentic_pipeline_tests" in u:
                            pr_link_tests = u
                        else:
                            pr_link_app = u

            results.append({
                **module_info,
                **doc_info,
                "Suite": suite.get("name"),
                "Test Class": classname,
                "Status": status,
                "Duration (s)": duration,
                "Message": message,
                "Suggested Fix": suggested_fix,
                "Short Summary": short_summary,
                "Jira Link": f'=HYPERLINK("{jira_link}","{jira_id}")' if (jira_id and jira_link) else (jira_id or ""),
                "Dev PR#": pr_link_app,
                "QA PR#": pr_link_tests,
                "Screenshot": screenshot_path,
                "Executed At": datetime.now().strftime("%Y_%m_%d-%H_%M_%S"),
                "IsHealed": "Yes" if is_healed else "",
            })
    return results

def _generate_excel_report_inline(root_log_dir, output_file=None, ai_state=None):

    logs_dir = Path( sd.log_folder_path )
    os.makedirs(
        str(logs_dir),
        exist_ok=True
    )
    xml_file = str(
        logs_dir /
        "test-results.xml"
    )

    if not os.path.exists(xml_file):
        (sd.logger or _logger).error("XML FILE NOT FOUND")
        return None

    if not output_file:
        date_folder = datetime.now().strftime(
            "%Y-%m-%d"
        )

        reports_dir = (
            Path(sd.reports_folder_path) /
            "excel" /
            date_folder
        )
        os.makedirs(
            reports_dir,
            exist_ok=True
        )
        
        time_stamp = datetime.now().strftime(
            "%Y-%m-%d_%H-%M-%S"
        )
        
        output_file = str(
            reports_dir /
            f"test-report_{time_stamp}.xlsx"
        )
        
    try:
        results = _parse_junit_xml(xml_file, root_log_dir, ai_state=ai_state)
    except Exception as e:
        print("\n FAILED TO PARSE XML")
        print(e)
        return None

    if not results:
        (sd.logger or _logger).warning("NO TEST RESULTS FOUND IN XML")
        _create_no_tests_report(output_file)
        return output_file

    (sd.logger or _logger).info(f"TOTAL TEST RESULTS PARSED: {len(results)}")
    try:
        _create_excel_report(results, output_file)
    except Exception as e:
        print("\n FAILED TO GENERATE EXCEL REPORT")
        print(e)
        return None

    (sd.logger or _logger).info(f"EXCEL REPORT GENERATED SUCCESSFULLY → {output_file}")
    return output_file


def generate_excel_from_json(json_path, output_file, ai_state=None):
    """Generate Excel report directly from a test_results JSON file.

    Used after a healed run where the JUnit XML only contains the original
    pre-healing collection errors while the JSON has all resolved results.
    """
    import json as _json
    from pathlib import Path as _Path

    json_path = _Path(json_path)
    if not json_path.exists():
        print(f"[excel_from_json] JSON not found: {json_path}")
        return None

    try:
        payload = _json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[excel_from_json] Failed to read JSON: {e}")
        return None

    results = []
    for entry in payload.get("results", []):
        test_id    = entry.get("test_id", "")
        test_name  = entry.get("test_name", "")
        status_raw = str(entry.get("status", "PASS")).upper()

        status_map = {
            "PASS": "PASSED", "PASSED": "PASSED",
            "FAIL": "FAILED", "FAILED": "FAILED",
            "ERROR": "ERROR", "SKIPPED": "SKIPPED",
        }
        status = status_map.get(status_raw, status_raw)

        # Build dot-notation classname from test_id
        # e.g. "test_framework/tests/integration/test_api_tasks.py::TestTaskAPI::test_create"
        classname = ""
        if "::" in test_id:
            parts = test_id.split("::")
            file_part = parts[0].replace("/", ".").replace(".py", "")
            classname = f"{file_part}.{parts[1]}" if len(parts) >= 3 else file_part
        else:
            classname = test_id.replace("/", ".").replace(".py", "")

        module_info = _extract_module_from_classname(classname)

        doc_info = {
            "Test Case Name":  entry.get("doc_test_case_name") or test_name,
            "Test Case ID":    entry.get("doc_test_case_id")   or "",
            "Module":          entry.get("doc_module")          or "",
            "Description":     entry.get("doc_description")     or "",
            "Steps":           entry.get("doc_steps")           or "",
            "Expected Output": entry.get("doc_expected_output") or "",
        }

        jira_id   = entry.get("jira_id",  "") or ""
        jira_url  = entry.get("jira_url", "") or ""
        is_healed = bool(entry.get("is_healed"))
        pr_url    = entry.get("pr_url",   "") or ""

        # Split PR URL into dev (app repo) and QA (tests repo) buckets
        pr_link_app = pr_link_tests = ""
        for u in str(pr_url).split(","):
            u = u.strip()
            if "/pull/" in u:
                num = u.rstrip("/").split("/")[-1]
                if "agentic_pipeline_tests" in u:
                    pr_link_tests = f'=HYPERLINK("{u}","PR #{num}")'
                else:
                    pr_link_app = f'=HYPERLINK("{u}","PR #{num}")'

        message       = entry.get("failure_reason", "") or ("Test passed successfully" if status == "PASSED" else "")
        suggested_fix = entry.get("ai_suggested_fix", "") or ""
        short_summary = entry.get("ai_short_summary",  "") or ""
        duration      = round(float(entry.get("duration", 0) or 0), 2)

        results.append({
            **module_info,
            **doc_info,
            "Suite":         "pytest",
            "Test Class":    classname,
            "Status":        status,
            "Duration (s)":  duration,
            "Message":       message,
            "Suggested Fix": suggested_fix,
            "Short Summary": short_summary,
            "Jira Link":     f'=HYPERLINK("{jira_url}","{jira_id}")' if (jira_id and jira_url) else (jira_id or ""),
            "Dev PR#":       pr_link_app,
            "QA PR#":        pr_link_tests,
            "Screenshot":    "",
            "Executed At":   datetime.now().strftime("%Y_%m_%d-%H_%M_%S"),
            "IsHealed":      "Yes" if is_healed else "",
        })

    if not results:
        print("[excel_from_json] No test entries found in JSON.")
        return None

    try:
        _create_excel_report(results, str(output_file))
        print(f"[excel_from_json] Excel generated → {output_file}")
        return str(output_file)
    except Exception as e:
        print(f"[excel_from_json] _create_excel_report failed: {e}")
        return None


def _extract_module_from_classname(
        classname: str
) -> dict:

    parts = classname.split(".") if classname else []

    feature = "Unknown"

    try:

        if "tests" in parts:

            tests_index = parts.index(
                "tests"
            )

            if (
                tests_index + 1
                < len(parts)
            ):

                feature = parts[
                    tests_index + 1
                ]

    except Exception:

        feature = "Unknown"

    return {

        "Platform": "Agentic Solution",

        "Feature": feature
    }

import re

def _parse_test_docstring(classname, test_name):
    import importlib
    import inspect
    import re

    result = {
        "Test Case Name": "",
        "Test Case ID": "",
        "Module": "",
        "Description": "",
        "Steps": "",
        "Expected Output": ""
    }

    try:
        parts = classname.split(".")
        _last = parts[-1] if parts else ""
        if _last and _last[0].isupper():
            module_name = ".".join(parts[:-1])
            class_name  = _last
        else:
            module_name = ".".join(parts)
            class_name  = None

        module = importlib.import_module(module_name) if module_name else None
        if not module:
            return result

        if class_name:
            container = getattr(module, class_name)
        else:
            container = module

        func = getattr(container, test_name, None)
        func_doc = inspect.cleandoc(func.__doc__ or "") if func else ""
        fallback_doc = inspect.cleandoc(container.__doc__ or "") if container else ""

        def extract_fields(doc_str):
            if not doc_str:
                return {}
            extracted = {}
            patterns = {
                "Test Case Name":   r"Test Case Name:\s*(.*?)(?:\n\s*\n|\Z)",
                "Module":           r"Module:\s*(.*?)(?:\n\s*\n|\Z)",
                "Test Case ID":     r"Test Case ID:\s*(.*?)(?:\n\s*\n|\Z)",
                "Description":      r"Description:\s*(.*?)(?:\n\s*\n|\Z)",
                "Steps":            r"Steps:\s*(.*?)(?:\n\s*\nExpected Output:|\n\s*\n|\Z)",
                "Expected Output":  r"Expected Output:\s*(.*)"
            }
            for key, pattern in patterns.items():
                match = re.search(pattern, doc_str, re.DOTALL)
                if match:
                    val = match.group(1).strip()
                    if val and val.upper() not in ("N/A", "NONE"):
                        extracted[key] = val
            return extracted

        func_fields = extract_fields(func_doc)
        fallback_fields = extract_fields(fallback_doc)

        def clean_val(v, default=""):
            if not v or str(v).strip().upper() in ("N/A", "NONE", ""):
                return default
            return str(v).strip()

        def clean_multiline(text):
            if not text:
                return ""
            return "\n".join(line.strip() for line in str(text).splitlines()).strip()

        for key in result.keys():
            val = func_fields.get(key) or fallback_fields.get(key)
            cleaned = clean_val(val, default="")
            if key in ("Description", "Steps", "Expected Output"):
                cleaned = clean_multiline(cleaned)
            result[key] = cleaned

        # Fallback values if completely empty
        if not result["Test Case Name"]:
            result["Test Case Name"] = test_name
        if not result["Module"]:
            result["Module"] = module_name.split(".")[-1] if module_name else "agentic_solution"
        if not result["Description"] and func_doc:
            result["Description"] = func_doc.strip()

        # Check for pytest.mark.testid decorator
        if func:
            marks = getattr(func, "pytestmark", [])
            for m in marks:
                if getattr(m, "name", None) == "testid" and getattr(m, "args", None):
                    result["Test Case ID"] = str(m.args[0]).strip()
                    break

        if "@testcase" in result["Test Case ID"]:
            result["Test Case ID"] = result["Test Case ID"].replace("@testcase", "").replace("ID:", "").replace("id:", "").strip("- ").strip()
        if not result["Test Case ID"]:
            import hashlib
            clean_test_name = test_name.split("[")[0]
            norm_id = f"{classname}.{clean_test_name}"
            h = hashlib.md5(norm_id.encode()).hexdigest()[:6].upper()
            result["Test Case ID"] = f"TC-{h}"

        return result

    except Exception as e:
        print("\nDOCSTRING PARSE ERROR:", e)
        return result


def _ensure_status_columns(df, statuses):
    for s in statuses:
        if s not in df.columns:
            df[s] = 0
    return df


def _apply_status_row_style(ws, row_idx, status_col_idx, jira_id_col_idx=None, pr_link_col_idx=None):
    from openpyxl.styles import PatternFill, Font

    status = ws.cell(row=row_idx, column=status_col_idx).value

    is_traceable = False
    if jira_id_col_idx:
        val = ws.cell(row=row_idx, column=jira_id_col_idx).value
        if val and str(val).strip() not in ("", "N/A", "None"):
            is_traceable = True
    if pr_link_col_idx:
        val = ws.cell(row=row_idx, column=pr_link_col_idx).value
        if val and str(val).strip() not in ("", "N/A", "None"):
            is_traceable = True

    if status == "PASSED":
        if is_traceable:
            # Healed / traceable test — soft violet
            fill, font = PatternFill("solid", "E8DFFF"), Font(color="4C1D95", bold=True)
        else:
            fill, font = PatternFill("solid", "C6EFCE"), Font(color="006100", bold=True)
    elif status in ("FAILED", "ERROR"):
        if is_traceable:
            # FAILED in the initial XML but has Jira/PR link → it was healed — show violet
            fill, font = PatternFill("solid", "E8DFFF"), Font(color="4C1D95", bold=True)
        else:
            fill, font = PatternFill("solid", "FFC7CE"), Font(color="9C0006", bold=True)
    elif status == "SKIPPED":
        fill, font = PatternFill("solid", "D9D9D9"), Font(bold=True)
    else:
        return

    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font


def _apply_border(ws, cell_range, thick=False):
    from openpyxl.styles import Border, Side

    style = "medium" if thick else "thin"
    border = Border(
        left=Side(style=style),
        right=Side(style=style),
        top=Side(style=style),
        bottom=Side(style=style),
    )

    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def _auto_adjust_column_width(ws):
    from openpyxl.utils import get_column_letter

    for idx, col in enumerate(ws.iter_cols(), start=1):
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 3, 60)


def _calculate_optimal_column_width(cell_values, min_width=10, max_width=100):
    max_length = min_width
    for value in cell_values:
        if value:
            lines = str(value).split('\n')
            longest_line = max(lines, key=len) if lines else ""
            max_length = max(max_length, len(longest_line))
    return min(max_length + 3, max_width)


def _calculate_optimal_row_height(row_values, column_widths, base_height=15, max_height=300):
    max_lines = 1
    for col_idx, value in enumerate(row_values, start=1):
        if value and isinstance(value, str):
            col_width = column_widths.get(col_idx, 15)
            lines = str(value).split('\n')
            total_lines = 0
            for line in lines:
                chars_per_line = int(col_width * 1.2)
                line_length = len(line)
                wrapped_lines = max(1, (line_length + chars_per_line - 1) // chars_per_line)
                total_lines += wrapped_lines
            max_lines = max(max_lines, total_lines)
    calculated_height = max_lines * base_height + 5
    return min(calculated_height, max_height)


def _apply_dynamic_column_widths(ws, df, header_row=1, data_start_row=2):
    from openpyxl.utils import get_column_letter

    column_constraints = {
        "Platform": {"min": 8, "max": 15},
        "Feature": {"min": 15, "max": 40},
        "Test Case Name": {"min": 30, "max": 80},
        "Test Case ID": {"min": 20, "max": 60},
        "Module": {"min": 20, "max": 60},
        "Description": {"min": 40, "max": 100},
        "Steps": {"min": 40, "max": 120},
        "Expected Output": {"min": 40, "max": 100},
        "Suite": {"min": 20, "max": 50},
        "Test Class": {"min": 30, "max": 80},
        "Status": {"min": 10, "max": 15},
        "Duration (s)": {"min": 10, "max": 15},
        "Message": {"min": 20, "max": 80},
        "Executed At": {"min": 15, "max": 25},
        "Suggested Fix": {"min": 40, "max": 120},
        "Short Summary": {"min": 40, "max": 120},
        "Jira Link": {"min": 15, "max": 25},
    }

    column_widths = {}

    for idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        cell_values = [cell.value for cell in ws[col_letter]]
        constraints = column_constraints.get(col_name, {"min": 10, "max": 80})
        optimal_width = _calculate_optimal_column_width(
            cell_values, min_width=constraints["min"], max_width=constraints["max"]
        )
        ws.column_dimensions[col_letter].width = optimal_width
        column_widths[idx] = optimal_width

    return column_widths


def _apply_dynamic_row_heights(ws, df, column_widths, data_start_row=2):
    for row_idx in range(data_start_row, ws.max_row + 1):
        row_values = [cell.value for cell in ws[row_idx]]
        optimal_height = _calculate_optimal_row_height(row_values, column_widths)
        ws.row_dimensions[row_idx].height = optimal_height


WEB_EXECUTION_TIME = 0.0

from openpyxl.styles import Alignment

def _align_dashboard_row(ws, row, label_align="left", value_align="right"):
    ws[f"A{row}"].alignment = Alignment(horizontal=label_align, vertical="center")
    ws[f"B{row}"].alignment = Alignment(horizontal=value_align, vertical="center")

REPORT_MODE = "LOCAL"

def _apply_screenshot_links(ws, df):

    if "Screenshot" not in df.columns:
        return

    report_mode = REPORT_MODE
    screenshot_idx = df.columns.get_loc("Screenshot") + 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=screenshot_idx)
        if not cell.value:
            continue
        cell.hyperlink = None
        filename = Path(cell.value).name
        if report_mode == "LOCAL":
            relative_path = Path("..") / ".." / "screenshots" / "milacron" / filename
        else:
            relative_path = Path("screenshots") / "milacron" / filename
        cell.value = "View Screenshot"
        cell.hyperlink = str(relative_path).replace("\\", "/")
        cell.font = Font(color="003366", underline="single", bold=True)


def _create_excel_report(test_results, output_file):
    import pandas as pd
    import sys
    try:
        import pandas.core.dtypes.inference as pandas_inference
        if not hasattr(pandas_inference, 'is_bool'):
            pandas_inference.is_bool = lambda x: isinstance(x, bool)
    except ImportError:
        import types
        m = types.ModuleType('pandas.core.dtypes.inference')
        m.is_bool = lambda x: isinstance(x, bool)
        sys.modules['pandas.core.dtypes.inference'] = m

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows

    def get_sort_key(row):
        status = str(row.get("Status", "")).upper()
        is_healed = bool(row.get("IsHealed", False))
        if status in ("FAILED", "ERROR"):
            return 0
        elif is_healed:
            return 1
        elif status == "PASSED":
            return 2
        elif status == "SKIPPED":
            return 3
        else:
            return 4

    test_results = sorted(test_results, key=get_sort_key)
    df = pd.DataFrame(test_results)
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    HEADER = PatternFill("solid", "366092")
    LIGHT = PatternFill("solid", "F2F2F2")
    RED = PatternFill("solid", "FFC7CE")
    GREY = PatternFill("solid", "D9D9D9")

    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True)
    value_font = Font(bold=True)

    ws_dash.column_dimensions["A"].width = 28
    ws_dash.column_dimensions["B"].width = 20
    ws_dash.column_dimensions["C"].width = 14
    ws_dash.column_dimensions["D"].width = 14
    ws_dash.column_dimensions["E"].width = 14
    ws_dash.column_dimensions["F"].width = 14

    current_row = 1

    ws_dash.merge_cells(f"A{current_row}:F{current_row}")
    cell = ws_dash[f"A{current_row}"]
    cell.value = "EXECUTION INFO"
    cell.fill = HEADER
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    _apply_border(ws_dash, f"A{current_row}:F{current_row}")

    current_row += 2

    exec_info = [
        ("Generated At", datetime.now().strftime("%Y_%m_%d-%H_%M_%S")),
        ("Web Version", os.getenv("WEB_APP_VERSION", "Feature Not Enabled")),
        ("Total Execution Time", str(timedelta(seconds=int(WEB_EXECUTION_TIME))))
    ]

    for label, value in exec_info:
        ws_dash[f"A{current_row}"] = label
        ws_dash[f"B{current_row}"] = value
        ws_dash[f"A{current_row}"].font = header_font
        ws_dash[f"B{current_row}"].font = value_font
        ws_dash[f"A{current_row}"].fill = LIGHT
        ws_dash[f"B{current_row}"].fill = LIGHT
        current_row += 1

    _apply_border(ws_dash, f"A{current_row-3}:B{current_row-1}")
    _align_dashboard_row(ws_dash, current_row)
    current_row += 2

    ws_dash.merge_cells(f"A{current_row}:F{current_row}")
    cell = ws_dash[f"A{current_row}"]
    cell.value = "BUILD HEALTH"
    cell.fill = HEADER
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    _apply_border(ws_dash, f"A{current_row}:F{current_row}")

    current_row += 2

    total = len(df)
    is_healed_series = df["IsHealed"].isin([True, 1, "Yes", "TRUE", "True"]) if "IsHealed" in df.columns else pd.Series([False]*len(df))
    healed_count = len(df[is_healed_series])
    passed = len(df[(df["Status"] == "PASSED") & (~is_healed_series)])
    failed = len(df[(df["Status"].isin(["FAILED", "ERROR"])) & (~is_healed_series)])
    skipped = len(df[df["Status"] == "SKIPPED"])
    pass_rate = ((passed + healed_count) / total * 100) if total else 0

    PURPLE = PatternFill("solid", "E8DFFF")

    build_rows = [
        ("Platform", "Agentic Solution"),
        ("Total Tests", total),
        ("Passed", passed),
        ("Healed Tests", healed_count),
        ("Failed / Errors", failed),
        ("Skipped", skipped),
        ("Pass Rate", f"{pass_rate:.2f}%")
    ]

    for label, value in build_rows:
        ws_dash[f"A{current_row}"] = label
        ws_dash[f"B{current_row}"] = value
        ws_dash[f"A{current_row}"].font = header_font
        ws_dash[f"B{current_row}"].font = value_font
        ws_dash[f"A{current_row}"].fill = LIGHT
        if label == "Failed / Errors" and failed:
            ws_dash[f"B{current_row}"].fill = RED
        elif label == "Skipped" and skipped:
            ws_dash[f"B{current_row}"].fill = GREY
        elif label == "Healed Tests" and healed_count:
            ws_dash[f"B{current_row}"].fill = PURPLE
        else:
            ws_dash[f"B{current_row}"].fill = LIGHT
        _align_dashboard_row(ws_dash, current_row)
        current_row += 1

    _apply_border(ws_dash, f"A{current_row-len(build_rows)}:B{current_row-1}")
    current_row += 2

    ws_dash.merge_cells(f"A{current_row}:F{current_row}")
    cell = ws_dash[f"A{current_row}"]
    cell.value = "TOP FAILING MODULES"
    cell.fill = HEADER
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    _apply_border(ws_dash, f"A{current_row}:F{current_row}")

    current_row += 2

    fail_modules = (
        df[(df["Status"].isin(["FAILED", "ERROR"])) & (~is_healed_series)]
        .groupby("Feature")
        .size()
        .reset_index(name="Failures")
        .sort_values("Failures", ascending=False)
        .head(5)
    )

    if fail_modules.empty:
        ws_dash.merge_cells(f"A{current_row}:F{current_row}")
        cell = ws_dash[f"A{current_row}"]
        cell.value = "🎉 No failing modules"
        cell.font = Font(bold=True)
        cell.fill = LIGHT
        cell.alignment = Alignment(horizontal="center")
        _apply_border(ws_dash, f"A{current_row}:F{current_row}")
    else:
        for _, row in fail_modules.iterrows():
            ws_dash[f"A{current_row}"] = row["Feature"]
            ws_dash[f"B{current_row}"] = row["Failures"]
            ws_dash[f"B{current_row}"].fill = RED
            current_row += 1
        _apply_border(ws_dash, f"A{current_row-len(fail_modules)}:B{current_row-1}")
        _align_dashboard_row(ws_dash, current_row)

    for platform in ["Agentic Solution"]:

        ws = wb.create_sheet(
            "Module Summary"
        )

        pdf = df[
            df["Platform"] == platform
        ]

        if pdf.empty:

            ws.append(
                [f"No {platform} tests executed"]
            )

            continue

        summary_rows = []

        for feature in pdf["Feature"].unique():

            fdf = pdf[
                pdf["Feature"] == feature
            ]

            total_tests = len(fdf)

            is_healed_fdf = fdf["IsHealed"].isin([True, 1, "Yes", "TRUE", "True"]) if "IsHealed" in fdf.columns else pd.Series([False]*len(fdf))
            healed = len(fdf[is_healed_fdf])

            passed = len(
                fdf[
                    (fdf["Status"] == "PASSED") & (~is_healed_fdf)
                ]
            )

            failed = len(
                fdf[
                    (fdf["Status"].isin(["FAILED", "ERROR"])) & (~is_healed_fdf)
                ]
            )

            skipped = len(
                fdf[
                    fdf["Status"] == "SKIPPED"
                ]
            )

            pass_rate = (
                round(
                    ((passed + healed) / total_tests) * 100,
                    2
                )
                if total_tests else 0
            )

            execution_duration = round(
                fdf["Duration (s)"]
                .astype(float)
                .sum(),
                2
            )

            # ai_failures variable removed — Healed count covers AI intervention

            summary_rows.append({

                "Feature": feature,

                "Total Tests": total_tests,

                "Passed": passed,

                "Healed": healed,

                "Failed": failed,

                "Skipped": skipped,

                "Pass Rate": f"{pass_rate}%",

                "Execution Duration (s)": execution_duration,
            })

        mdf = pd.DataFrame(
            summary_rows
        )

        ws.append(
            list(mdf.columns)
        )

        for cell in ws[1]:

            cell.fill = HEADER

            cell.font = Font(
                color="FFFFFF",
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        for row_idx, row in enumerate(
                dataframe_to_rows(
                    mdf,
                    index=False,
                    header=False
                ),
                start=2
        ):

            ws.append(row)

            pass_rate_value = str(
                ws.cell(
                    row=row_idx,
                    column=7
                ).value
            ).replace("%", "")

            try:

                pass_rate_value = float(
                    pass_rate_value
                )

            except Exception:

                pass_rate_value = 0

            for cell in ws[row_idx]:

                #cell.fill = fill

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = ws.dimensions

        _auto_adjust_column_width(ws)

    ws_details = wb.create_sheet("Test Details")
    ws_details.append(list(df.columns))

    for cell in ws_details[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_idx = df.columns.get_loc("Status") + 1
    screenshot_idx = df.columns.get_loc("Screenshot") + 1
    desc_idx = df.columns.get_loc("Description") + 1 if "Description" in df.columns else None
    steps_idx = df.columns.get_loc("Steps") + 1 if "Steps" in df.columns else None
    output_idx = df.columns.get_loc("Expected Output") + 1 if "Expected Output" in df.columns else None
    module_idx = df.columns.get_loc("Module") + 1
    tcid_idx = df.columns.get_loc("Test Case ID") + 1
    fix_idx = df.columns.get_loc("Suggested Fix") + 1 if "Suggested Fix" in df.columns else None
    summary_idx = df.columns.get_loc("Short Summary") + 1 if "Short Summary" in df.columns else None
    jira_link_idx = df.columns.get_loc("Jira Link") + 1 if "Jira Link" in df.columns else None
    msg_idx = df.columns.get_loc("Message") + 1 if "Message" in df.columns else None

    pr_app_idx = df.columns.get_loc("Dev PR#") + 1 if "Dev PR#" in df.columns else None
    pr_tests_idx = df.columns.get_loc("QA PR#") + 1 if "QA PR#" in df.columns else None

    for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        ws_details.append(row)
        # Pass pr_app_idx or pr_tests_idx as the pr_link_col_idx to color healed rows violet
        _apply_status_row_style(ws_details, r, status_idx, jira_link_idx, pr_app_idx or pr_tests_idx)
        
        # Standardize all columns to wrap text and align top-left
        for col_idx in range(1, ws_details.max_column + 1):
            ws_details.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            
        # Jira Link — single-click HYPERLINK formula formatting
        if jira_link_idx:
            cell = ws_details.cell(row=r, column=jira_link_idx)
            val = cell.value
            if val and str(val).startswith("=HYPERLINK"):
                cell.font = Font(color="0563C1", underline="single", bold=True)

        # Dev PR# — clickable HYPERLINK
        if pr_app_idx:
            cell = ws_details.cell(row=r, column=pr_app_idx)
            val = cell.value
            if val and str(val).strip() and not str(val).startswith("="):
                safe_url = str(val).replace('"', '').strip()
                if "/pull/" in safe_url:
                    pr_num = safe_url.rstrip('/').split('/')[-1]
                    pr_text = f"PR #{pr_num}" if (pr_num and pr_num.isdigit()) else "PR"
                    cell.value = f'=HYPERLINK("{safe_url}","{pr_text}")'
                    cell.font = Font(color="7C3AED", underline="single", bold=True)

        # QA PR# — clickable HYPERLINK
        if pr_tests_idx:
            cell = ws_details.cell(row=r, column=pr_tests_idx)
            val = cell.value
            if val and str(val).strip() and not str(val).startswith("="):
                safe_url = str(val).replace('"', '').strip()
                if "/pull/" in safe_url:
                    pr_num = safe_url.rstrip('/').split('/')[-1]
                    pr_text = f"PR #{pr_num}" if (pr_num and pr_num.isdigit()) else "PR"
                    cell.value = f'=HYPERLINK("{safe_url}","{pr_text}")'
                    cell.font = Font(color="7C3AED", underline="single", bold=True)

    ws_details.freeze_panes = "D2"
    ws_details.auto_filter.ref = ws_details.dimensions
    column_widths = _apply_dynamic_column_widths(ws_details, df)
    _apply_dynamic_row_heights(ws_details, df, column_widths)
    _apply_border(ws_details, ws_details.dimensions)
    _align_dashboard_row(ws_dash, current_row)

    ws_failed = wb.create_sheet("Failed Tests")
    # Exclude healed tests — only show genuine (non-healed) failures
    _is_healed_col = df["IsHealed"].isin([True, 1, "Yes", "TRUE", "True"]) if "IsHealed" in df.columns else pd.Series([False]*len(df))
    fdf = df[(df["Status"].isin(["FAILED", "ERROR"])) & (~_is_healed_col)]

    if fdf.empty:
        ws_failed.append(["No failed tests"])
    else:
        ws_failed.append(list(fdf.columns))
        for cell in ws_failed[1]:
            cell.fill = HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        status_idx = fdf.columns.get_loc("Status") + 1
        screenshot_idx = fdf.columns.get_loc("Screenshot") + 1
        desc_idx = fdf.columns.get_loc("Description") + 1 if "Description" in fdf.columns else None
        steps_idx = fdf.columns.get_loc("Steps") + 1 if "Steps" in fdf.columns else None
        output_idx = fdf.columns.get_loc("Expected Output") + 1 if "Expected Output" in fdf.columns else None
        fix_idx = fdf.columns.get_loc("Suggested Fix") + 1 if "Suggested Fix" in fdf.columns else None
        summary_idx = fdf.columns.get_loc("Short Summary") + 1 if "Short Summary" in fdf.columns else None
        jira_link_idx_f = fdf.columns.get_loc("Jira Link") + 1 if "Jira Link" in fdf.columns else None
        msg_idx_f = fdf.columns.get_loc("Message") + 1 if "Message" in fdf.columns else None

        for r, row in enumerate(dataframe_to_rows(fdf, index=False, header=False), start=2):
            ws_failed.append(row)
            _apply_status_row_style(ws_failed, r, status_idx)
            
            # Standardize all columns to wrap text and align top-left
            for col_idx in range(1, ws_failed.max_column + 1):
                ws_failed.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                
            # Jira Link — single-click HYPERLINK formula formatting
            if jira_link_idx_f:
                cell = ws_failed.cell(row=r, column=jira_link_idx_f)
                val = cell.value
                if val and str(val).startswith("=HYPERLINK"):
                    cell.font = Font(color="0563C1", underline="single", bold=True)

        ws_failed.freeze_panes = "D2"
        ws_failed.auto_filter.ref = ws_failed.dimensions
        column_widths_failed = _apply_dynamic_column_widths(ws_failed, fdf)
        _apply_dynamic_row_heights(ws_failed, fdf, column_widths_failed)
        _apply_border(ws_failed, ws_failed.dimensions)

    ws_healed = wb.create_sheet("Healed Tests")
    PURPLE_HEADER = PatternFill("solid", "5B21B6")   # deep violet header
    PURPLE_ROW    = PatternFill("solid", "EDE9FE")   # light lavender row bg
    _is_healed_col2 = df["IsHealed"].isin([True, 1, "Yes", "TRUE", "True"]) if "IsHealed" in df.columns else pd.Series([False]*len(df))
    hdf = df[_is_healed_col2]

    if hdf.empty:
        ws_healed.append(["No healed tests in this run"])
        ws_healed.cell(row=1, column=1).font = Font(bold=True, italic=True)
    else:
        ws_healed.append(list(hdf.columns))
        for cell in ws_healed[1]:
            cell.fill = PURPLE_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        status_idx_h   = hdf.columns.get_loc("Status") + 1
        jira_link_idx_h = hdf.columns.get_loc("Jira Link") + 1 if "Jira Link" in hdf.columns else None
        pr_app_idx_h   = hdf.columns.get_loc("Dev PR#") + 1 if "Dev PR#" in hdf.columns else None
        pr_tests_idx_h = hdf.columns.get_loc("QA PR#") + 1 if "QA PR#" in hdf.columns else None

        for r, row in enumerate(dataframe_to_rows(hdf, index=False, header=False), start=2):
            ws_healed.append(row)

            # Purple row background
            for col_idx in range(1, ws_healed.max_column + 1):
                cell = ws_healed.cell(row=r, column=col_idx)
                cell.fill = PURPLE_ROW
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            # Jira Link — clickable hyperlink
            if jira_link_idx_h:
                cell = ws_healed.cell(row=r, column=jira_link_idx_h)
                val = cell.value
                if val and str(val).startswith("=HYPERLINK"):
                    cell.font = Font(color="0563C1", underline="single", bold=True)
                elif val and str(val).strip() and not str(val).startswith("="):
                    # plain Jira ID with no URL — still style it
                    cell.font = Font(color="0563C1", bold=True)

            # Dev PR# — clickable HYPERLINK
            if pr_app_idx_h:
                cell = ws_healed.cell(row=r, column=pr_app_idx_h)
                val = cell.value
                if val and str(val).strip() and not str(val).startswith("="):
                    safe_url = str(val).replace('"', '').strip()
                    if "/pull/" in safe_url:
                        pr_num = safe_url.rstrip('/').split('/')[-1]
                        pr_text = f"PR #{pr_num}" if (pr_num and pr_num.isdigit()) else "PR"
                        cell.value = f'=HYPERLINK("{safe_url}","{pr_text}")'
                        cell.font = Font(color="7C3AED", underline="single", bold=True)

            # QA PR# — clickable HYPERLINK
            if pr_tests_idx_h:
                cell = ws_healed.cell(row=r, column=pr_tests_idx_h)
                val = cell.value
                if val and str(val).strip() and not str(val).startswith("="):
                    safe_url = str(val).replace('"', '').strip()
                    if "/pull/" in safe_url:
                        pr_num = safe_url.rstrip('/').split('/')[-1]
                        pr_text = f"PR #{pr_num}" if (pr_num and pr_num.isdigit()) else "PR"
                        cell.value = f'=HYPERLINK("{safe_url}","{pr_text}")'
                        cell.font = Font(color="7C3AED", underline="single", bold=True)

        ws_healed.freeze_panes = "D2"
        ws_healed.auto_filter.ref = ws_healed.dimensions
        column_widths_healed = _apply_dynamic_column_widths(ws_healed, hdf)
        _apply_dynamic_row_heights(ws_healed, hdf, column_widths_healed)
        _apply_border(ws_healed, ws_healed.dimensions)

    ws_skipped = wb.create_sheet("Skipped Tests")
    sdf = df[df["Status"] == "SKIPPED"]

    if sdf.empty:
        ws_skipped.append(["No skipped tests"])
    else:
        ws_skipped.append(list(sdf.columns))
        for cell in ws_skipped[1]:
            cell.fill = HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        status_idx = sdf.columns.get_loc("Status") + 1
        desc_idx = sdf.columns.get_loc("Description") + 1 if "Description" in sdf.columns else None
        steps_idx = sdf.columns.get_loc("Steps") + 1 if "Steps" in sdf.columns else None
        output_idx = sdf.columns.get_loc("Expected Output") + 1 if "Expected Output" in sdf.columns else None
        fix_idx = sdf.columns.get_loc("Suggested Fix") + 1 if "Suggested Fix" in sdf.columns else None
        summary_idx = sdf.columns.get_loc("Short Summary") + 1 if "Short Summary" in sdf.columns else None

        for r, row in enumerate(dataframe_to_rows(sdf, index=False, header=False), start=2):
            ws_skipped.append(row)
            _apply_status_row_style(ws_skipped, r, status_idx)
            
            # Standardize all columns to wrap text and align top-left
            for col_idx in range(1, ws_skipped.max_column + 1):
                ws_skipped.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        _apply_screenshot_links(ws_skipped, sdf)
        ws_skipped.freeze_panes = "D2"
        ws_skipped.auto_filter.ref = ws_skipped.dimensions
        column_widths_skipped = _apply_dynamic_column_widths(ws_skipped, sdf)
        _apply_dynamic_row_heights(ws_skipped, sdf, column_widths_skipped)
        _apply_border(ws_skipped, ws_skipped.dimensions)

    wb.save(output_file)


def _create_no_tests_report(output_file):
    from openpyxl import Workbook
    wb = Workbook()
    wb.active["A1"] = "No tests were executed"
    wb.save(output_file)