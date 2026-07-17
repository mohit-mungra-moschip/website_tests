#!/usr/bin/env python3
"""
utils/report_utils/update_reports_pr.py

Post-run script to dynamically update JSON, HTML, and Excel reports with
the actual GitHub Pull Request URL after the PR is successfully created.
"""

import argparse
import json
import os
import sys
import glob
from pathlib import Path

# Add root folder to sys.path to import conftest
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))
try:
    import conftest
except ImportError:
    conftest = None

def parse_args():
    parser = argparse.ArgumentParser(description="Update test reports with actual PR URL.")
    parser.add_argument("--pr-url", required=True, help="Actual GitHub PR URL.")
    parser.add_argument("--type", choices=["APP_HEAL", "TEST_HEAL", "MIXED"], default="APP_HEAL", help="Healing type.")
    parser.add_argument("--run-id", help="Pipeline Run ID.")
    return parser.parse_args()

def find_latest_report(directory, pattern):
    files = glob.glob(os.path.join(directory, pattern))
    if not files:
        return None
    # Sort by modification time desc
    files.sort(key=os.path.getmtime, reverse=True)
    return Path(files[0])

def get_base_dir():
    cwd = Path(".").resolve()
    
    # 1. In GitHub Actions, tests repo is checked out under "test_framework" subdirectory
    if cwd.name == "test_framework" or os.path.basename(os.getcwd()) == "test_framework":
        parent = cwd.parent if cwd.name == "test_framework" else Path("..").resolve()
        if (parent / "reports").exists():
            return parent.resolve()

    # 2. Local dev where agentic_pipeline_tests is a sibling of agentic_pipeline
    sibling_app = cwd.parent / "agentic_pipeline"
    if sibling_app.exists() and (sibling_app / "reports").exists():
        if cwd.name in ("agentic_pipeline_tests", "test_framework") or "tests" in cwd.name:
            return sibling_app.resolve()
            
    # 3. Default to current directory if reports folder exists here and we are not in tests folder
    if (cwd / "reports").exists() and not (cwd.name in ("agentic_pipeline_tests", "test_framework") or "tests" in cwd.name):
        return cwd
        
    # 4. Fallback to parent
    if (cwd.parent / "reports").exists():
        return cwd.parent.resolve()
        
    # 5. Fallback to sibling if everything else fails
    if sibling_app.exists():
        return sibling_app.resolve()
        
    return cwd

def update_json_and_html(run_id, pr_url):
    print(f"[*] Locating JSON/HTML reports for Run ID: {run_id or 'latest'}...")
    base_dir = get_base_dir()

    json_dir = base_dir / "reports/json"
    html_dir = base_dir / "reports/html"

    target_files = []
    if run_id:
        # 1. First, check if run_id matches the timestamp in filename directly
        for suffix in ["", "_healed", "_full_rerun"]:
            jp = json_dir / f"test_results_{run_id}{suffix}.json"
            hp = html_dir / f"test_results_{run_id}{suffix}.html"
            if jp.exists():
                target_files.append((jp, hp))
        
        # 2. If not found, scan all test_results_*.json files and match run_id inside the payload
        if not target_files:
            for jf in json_dir.glob("test_results_*.json"):
                try:
                    payload = json.loads(jf.read_text(encoding="utf-8"))
                    file_run_id = str(payload.get("run_id") or "")
                    if file_run_id == str(run_id) or file_run_id == f"{run_id}_healed" or file_run_id.startswith(str(run_id)):
                        stem = jf.stem
                        base_stem = stem.replace("_healed", "").replace("_full_rerun", "")
                        for suffix in ["", "_healed", "_full_rerun"]:
                            jp = json_dir / f"{base_stem}{suffix}.json"
                            hp = html_dir / f"{base_stem}{suffix}.html"
                            if jp.exists() and (jp, hp) not in target_files:
                                target_files.append((jp, hp))
                except Exception:
                    pass
                    
    # 3. Fallback to latest run if no files matched or no run_id provided
    if not target_files:
        json_path = find_latest_report(str(json_dir), "test_results_*.json")
        if json_path:
            stem = json_path.stem
            clean_run_id = stem.replace("test_results_", "").replace("_full_rerun", "").replace("_healed", "")
            for suffix in ["", "_healed", "_full_rerun"]:
                jp = json_dir / f"test_results_{clean_run_id}{suffix}.json"
                hp = html_dir / f"test_results_{clean_run_id}{suffix}.html"
                if jp.exists():
                    target_files.append((jp, hp))

    existing_targets = [(jp, hp) for jp, hp in target_files if jp.exists()]

    if not existing_targets:
        print(f"[!] JSON report files not found. Skipping JSON/HTML update.")
        return False

    for json_path, html_path in existing_targets:
        print(f"[*] Patching JSON report: {json_path}")
        try:
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception as e:
            print(f"[x] Error reading JSON: {e}")
            continue

        updated = False
        for result in payload.get("results", []):
            curr_pr = result.get("pr_url", "")
            status = result.get("status", "")
            is_healed = result.get("is_healed") or (status in ("PASS", "PASSED") and result.get("jira_id")) or "ai-fix" in str(curr_pr) or "/pull/" in str(curr_pr)
            if is_healed:
                existing_pr = result.get("pr_url", "")
                if existing_pr:
                    existing_list = [x.strip() for x in str(existing_pr).split(",") if x.strip()]
                    # Filter out placeholder/tree links
                    existing_list = [x for x in existing_list if "/tree/" not in x]
                    if pr_url not in existing_list:
                        existing_list.append(pr_url)
                        result["pr_url"] = ",".join(existing_list)
                        updated = True
                else:
                    result["pr_url"] = pr_url
                    updated = True

        if updated:
            results_list = payload.get("results", [])
            total = len(results_list)
            passed = sum(1 for r in results_list if r.get("status") in ("PASS", "PASSED") and not r.get("pr_url"))
            failed = sum(1 for r in results_list if r.get("status") in ("FAIL", "FAILED", "ERROR") and not r.get("pr_url"))
            skipped = sum(1 for r in results_list if r.get("status") == "SKIPPED")
            healed = sum(1 for r in results_list if r.get("pr_url"))
            
            payload["summary"] = {
                "total": total,
                "passed": passed,
                "failed": failed,
                "healed": healed,
                "skipped": skipped,
                "success_rate": round(((passed + healed) / total * 100) if total > 0 else 0.0, 2)
            }
            json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"[+] JSON report updated successfully.")

            print(f"[*] Regenerating/patching HTML report: {html_path}")
            try:
                if conftest is not None:
                    html_content = conftest._build_html(payload, json_path.name)
                    html_path.write_text(html_content, encoding="utf-8")
                    print(f"[+] HTML report regenerated successfully using conftest.")
                else:
                    print(f"[*] conftest not imported (missing pytest). Attempting direct patch of existing HTML file...")
                    if html_path.exists():
                        content = html_path.read_text(encoding="utf-8")
                        import re
                        pattern = r'https://github\.com/(softnauticsgithub|mohit-mungra-moschip)/[a-zA-Z0-9_-]+/tree/ai-fix/[a-zA-Z0-9_-]+'
                        new_content = re.sub(pattern, pr_url, content)
                        html_path.write_text(new_content, encoding="utf-8")
                        print(f"[+] HTML report patched directly via regex replacement.")
                    else:
                        print(f"[!] HTML report does not exist. Cannot patch.")
            except Exception as e:
                print(f"[x] Error updating HTML report: {e}")
        else:
            print(f"[-] No healed/traceable test cases found in JSON report: {json_path.name}")

    return True

def extract_urls_from_cell(val):
    if not val:
        return []
    val_str = str(val).strip()
    if not val_str:
        return []
    
    import re
    urls = []
    found = re.findall(r'https?://[^\s"\',)]+', val_str)
    for u in found:
        u_clean = u.strip()
        if u_clean and u_clean not in urls:
            urls.append(u_clean)
            
    if not urls:
        for chunk in re.split(r'[,\n]', val_str):
            chunk_clean = chunk.strip()
            if chunk_clean.startswith("http"):
                if chunk_clean not in urls:
                    urls.append(chunk_clean)
    return urls

def update_excel_report(pr_url, run_id=None):
    base_dir = get_base_dir()

    reports_dir = base_dir / "reports"
    excel_path = None
    
    # 1. Direct match if run_id is a timestamp format
    if run_id:
        ep = reports_dir / f"test_results_{run_id}.xlsx"
        if ep.exists():
            excel_path = ep
            
    # 2. Match by searching JSON files
    if not excel_path and run_id:
        json_dir = reports_dir / "json"
        for jf in json_dir.glob("test_results_*.json"):
            try:
                payload = json.loads(jf.read_text(encoding="utf-8"))
                file_run_id = str(payload.get("run_id") or "")
                if file_run_id == str(run_id) or file_run_id == f"{run_id}_healed" or file_run_id.startswith(str(run_id)):
                    stem = jf.stem.replace("_healed", "").replace("_full_rerun", "")
                    ep = reports_dir / f"{stem}.xlsx"
                    if ep.exists():
                        excel_path = ep
                        break
            except Exception:
                pass

    # 3. Fallback to latest test_results_*.xlsx file
    if not excel_path:
        excel_files = list(reports_dir.glob("test_results_*.xlsx"))
        if excel_files:
            excel_files.sort(key=os.path.getmtime, reverse=True)
            excel_path = excel_files[0]
            
    # 4. Fallback to default test_results.xlsx
    if not excel_path:
        ep = reports_dir / "test_results.xlsx"
        if ep.exists():
            excel_path = ep

    if not excel_path or not excel_path.exists():
        print(f"[!] Excel report not found. Skipping Excel update.")
        return False

    print(f"[*] Locating and patching Excel report: {excel_path}")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[x] openpyxl not installed. Cannot update Excel report.")
        return False

    # Load JSON reports for mapping test cases to their Jira/PR info
    json_map = {}
    try:
        reports_dir = base_dir / "reports"
        json_dir = reports_dir / "json"
        json_files = []
        if run_id:
            json_files.extend(list(reports_dir.glob(f"test_results_{run_id}.json")))
            if json_dir.exists():
                json_files.extend(list(json_dir.glob(f"test_results_{run_id}.json")))
            
            if not json_files:
                all_candidates = []
                if json_dir.exists():
                    all_candidates.extend(list(json_dir.glob("test_results_*.json")))
                all_candidates.extend(list(reports_dir.glob("test_results_*.json")))
                for jf in all_candidates:
                    try:
                        payload = json.loads(jf.read_text(encoding="utf-8"))
                        file_run_id = str(payload.get("run_id") or "")
                        if (file_run_id == str(run_id) or file_run_id == f"{run_id}_healed" or file_run_id.startswith(str(run_id))) and jf not in json_files:
                            json_files.append(jf)
                    except Exception:
                        pass
                
        for jf in json_files:
            try:
                payload = json.loads(jf.read_text(encoding="utf-8"))
                for res in payload.get("results", []):
                    t_id = res.get("test_id")
                    t_name = res.get("test_name")
                    doc_tc_id = res.get("doc_test_case_id")
                    doc_tc_name = res.get("doc_test_case_name")
                    if t_id:
                        json_map[t_id] = res
                    if t_name:
                        json_map[t_name] = res
                    if doc_tc_id:
                        json_map[doc_tc_id] = res
                    if doc_tc_name:
                        json_map[doc_tc_name] = res
            except Exception as je:
                pass
    except Exception as e:
        print(f"[!] Warning: failed to load JSON reports for Excel mapping: {e}")

    try:
        wb = load_workbook(str(excel_path))
        if "Test Details" not in wb.sheetnames:
            print("[!] 'Test Details' sheet not found in Excel. Skipping Excel update.")
            return False

        ws = wb["Test Details"]
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        
        try:
            status_idx = headers.index("Status") + 1
        except ValueError:
            print("[x] 'Status' column not found in Excel sheet.")
            return False

        jira_id_idx = None
        if "Jira Link" in headers:
            jira_id_idx = headers.index("Jira Link") + 1

        ishealed_idx = None
        if "IsHealed" in headers:
            ishealed_idx = headers.index("IsHealed") + 1

        pr_app_idx = None
        if "Dev PR#" in headers:
            pr_app_idx = headers.index("Dev PR#") + 1

        pr_tests_idx = None
        if "QA PR#" in headers:
            pr_tests_idx = headers.index("QA PR#") + 1

        tcid_idx = None
        if "Test Case ID" in headers:
            tcid_idx = headers.index("Test Case ID") + 1

        name_idx = None
        if "Test Case Name" in headers:
            name_idx = headers.index("Test Case Name") + 1

        print(f"[*] 'Status' col: {status_idx}, 'Jira Link' col: {jira_id_idx}, 'Dev PR#' col: {pr_app_idx}, 'QA PR#' col: {pr_tests_idx}")

        updated_count = 0
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=status_idx).value
            
            tc_id = ws.cell(row=r, column=tcid_idx).value if tcid_idx else None
            tc_name = ws.cell(row=r, column=name_idx).value if name_idx else None

            res = None
            if tc_id and tc_id in json_map:
                res = json_map[tc_id]
            elif tc_name and tc_name in json_map:
                res = json_map[tc_name]

            # 1. Update/populate Jira Link column if it is empty/N/A
            jira_id = ""
            jira_url = ""
            if res:
                jira_id = res.get("jira_id")
                jira_url = res.get("jira_url")

            if jira_id_idx and jira_id:
                jira_cell = ws.cell(row=r, column=jira_id_idx)
                jira_val = jira_cell.value
                if not jira_val or str(jira_val).strip() in ("", "N/A", "None"):
                    if jira_url:
                        jira_cell.value = f'=HYPERLINK("{jira_url}","{jira_id}")'
                    else:
                        jira_cell.value = jira_id
                    jira_cell.font = Font(color="0563C1", underline="single", bold=True)

            has_jira = False
            if jira_id_idx:
                jira_val = ws.cell(row=r, column=jira_id_idx).value
                if jira_val and str(jira_val).strip() not in ("", "N/A", "None"):
                    has_jira = True

            # Determine if this row is healed
            is_healed = False
            if ishealed_idx:
                is_h_val = ws.cell(row=r, column=ishealed_idx).value
                if is_h_val in (True, 1, "TRUE", "1", "True", "Yes", "yes", "YES"):
                    is_healed = True
            
            if not is_healed:
                is_healed = (status == "PASSED" and has_jira) or (res and res.get("is_healed"))

            if is_healed and ishealed_idx:
                ws.cell(row=r, column=ishealed_idx).value = "Yes"

            if is_healed:
                # Update the new incoming PR URL to the appropriate column
                is_test_pr = "agentic_pipeline_tests" in pr_url
                target_col_idx = pr_tests_idx if is_test_pr else pr_app_idx
                if target_col_idx:
                    pr_cell = ws.cell(row=r, column=target_col_idx)
                    pr_cell.value = pr_url

                # Also populate other column if we have a mapped pr_url in JSON
                if res and res.get("pr_url"):
                    for u in str(res.get("pr_url")).split(","):
                        u_clean = u.strip()
                        if u_clean and u_clean.startswith("http") and "/pull/" in u_clean:
                            u_is_test = "agentic_pipeline_tests" in u_clean
                            other_col_idx = pr_tests_idx if u_is_test else pr_app_idx
                            if other_col_idx:
                                other_cell = ws.cell(row=r, column=other_col_idx)
                                if not other_cell.value or str(other_cell.value).strip() in ("", "None", "N/A"):
                                    other_cell.value = u_clean

                # Now style both columns if they have pull requests!
                row_updated = False
                for col_idx in [pr_app_idx, pr_tests_idx]:
                    if col_idx:
                        cell = ws.cell(row=r, column=col_idx)
                        val = cell.value
                        if val and str(val).strip() and not str(val).startswith("="):
                            safe_url = str(val).replace('"', '').strip()
                            if "/pull/" in safe_url:
                                pr_num = safe_url.rstrip('/').split('/')[-1]
                                pr_text = f"PR #{pr_num}" if (pr_num and pr_num.isdigit()) else "PR"
                                cell.value = f'=HYPERLINK("{safe_url}","{pr_text}")'
                                cell.font = Font(color="7C3AED", underline="single", bold=True)
                                row_updated = True

                # Apply healed style (soft violet background + deep purple font)
                fill = PatternFill("solid", "E8DFFF")
                font = Font(color="4C1D95", bold=True)
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.fill = fill
                    if c != pr_app_idx and c != pr_tests_idx and c != jira_id_idx:
                        cell.font = font
                
                if row_updated:
                    updated_count += 1

        if updated_count > 0:
            print(f"[+] Test Details sheet: patched {updated_count} rows.")
        else:
            print(f"[-] No healed rows found in Test Details sheet to update.")

        # ── Also patch the Healed Tests sheet (same PR URL logic) ──────────
        healed_count_h = 0
        if "Healed Tests" in wb.sheetnames:
            ws_h = wb["Healed Tests"]
            h_headers = [ws_h.cell(row=1, column=col).value for col in range(1, ws_h.max_column + 1)]
            h_pr_app_idx = (h_headers.index("Dev PR#") + 1) if "Dev PR#" in h_headers else None
            h_pr_tests_idx = (h_headers.index("QA PR#") + 1) if "QA PR#" in h_headers else None
            h_jira_idx = (h_headers.index("Jira Link") + 1) if "Jira Link" in h_headers else None
            h_tcid_idx = (h_headers.index("Test Case ID") + 1) if "Test Case ID" in h_headers else None
            h_name_idx = (h_headers.index("Test Case Name") + 1) if "Test Case Name" in h_headers else None

            for r in range(2, ws_h.max_row + 1):
                tc_id_h = ws_h.cell(row=r, column=h_tcid_idx).value if h_tcid_idx else None
                tc_name_h = ws_h.cell(row=r, column=h_name_idx).value if h_name_idx else None

                # Populate Jira Link if missing
                res_h = json_map.get(tc_id_h) or json_map.get(tc_name_h)
                if h_jira_idx and res_h:
                    jira_cell_h = ws_h.cell(row=r, column=h_jira_idx)
                    jira_val_h = jira_cell_h.value
                    if not jira_val_h or str(jira_val_h).strip() in ("", "N/A", "None"):
                        jira_id_h = res_h.get("jira_id", "")
                        jira_url_h = res_h.get("jira_url", "")
                        if jira_id_h and jira_url_h:
                            jira_cell_h.value = f'=HYPERLINK("{jira_url_h}","{jira_id_h}")'
                            jira_cell_h.font = Font(color="0563C1", underline="single", bold=True)
                        elif jira_id_h:
                            jira_cell_h.value = jira_id_h
                            jira_cell_h.font = Font(color="0563C1", bold=True)

                # Update the new incoming PR URL to the appropriate column
                is_test_pr = "agentic_pipeline_tests" in pr_url
                target_col_h = h_pr_tests_idx if is_test_pr else h_pr_app_idx
                if target_col_h:
                    pr_cell_h = ws_h.cell(row=r, column=target_col_h)
                    pr_cell_h.value = pr_url

                # Also populate other column if we have a mapped pr_url in JSON
                if res_h and res_h.get("pr_url"):
                    for u in str(res_h.get("pr_url")).split(","):
                        u_clean = u.strip()
                        if u_clean and u_clean.startswith("http") and "/pull/" in u_clean:
                            u_is_test = "agentic_pipeline_tests" in u_clean
                            other_col_h = h_pr_tests_idx if u_is_test else h_pr_app_idx
                            if other_col_h:
                                other_cell_h = ws_h.cell(row=r, column=other_col_h)
                                if not other_cell_h.value or str(other_cell_h.value).strip() in ("", "None", "N/A"):
                                    other_cell_h.value = u_clean

                # Now style both columns if they have pull requests!
                row_updated = False
                for col_idx in [h_pr_app_idx, h_pr_tests_idx]:
                    if col_idx:
                        cell_h = ws_h.cell(row=r, column=col_idx)
                        val_h = cell_h.value
                        if val_h and str(val_h).strip() and not str(val_h).startswith("="):
                            safe_url_h = str(val_h).replace('"', '').strip()
                            if "/pull/" in safe_url_h:
                                pr_num_h = safe_url_h.rstrip('/').split('/')[-1]
                                pr_txt = f"PR #{pr_num_h}" if (pr_num_h and pr_num_h.isdigit()) else "PR"
                                cell_h.value = f'=HYPERLINK("{safe_url_h}","{pr_txt}")'
                                cell_h.font = Font(color="7C3AED", underline="single", bold=True)
                                row_updated = True
                
                if row_updated:
                    healed_count_h += 1

            print(f"[+] Healed Tests sheet: patched {healed_count_h} rows.")
        else:
            print("[!] 'Healed Tests' sheet not found — skipping.")

        if updated_count > 0 or healed_count_h > 0:
            wb.save(str(excel_path))
            print(f"[+] Excel report saved. ({updated_count} Test Details + {healed_count_h} Healed Tests rows updated)")
        else:
            print(f"[-] No rows updated in Excel report.")

    except Exception as e:
        print(f"[x] Error updating Excel report: {e}")
        return False

    return True

def main():
    args = parse_args()
    run_id = args.run_id or os.environ.get("REGRESSION_RUN_ID")
    
    success_json = update_json_and_html(run_id, args.pr_url)
    success_excel = update_excel_report(args.pr_url, run_id)
    
    if success_json or success_excel:
        print("[+] All reports updated successfully.")
    else:
        print("[-] Report update finished with no changes.")

    # Send pipeline report email after reports have been patched with the PR URL
    try:
        from utils.mailer import send_pipeline_report
        base_dir = get_base_dir()
        summary_path = base_dir / "reports/ai_summary.json"
        if summary_path.exists():
            print(f"[*] Loading summary to send report email: {summary_path}")
            state = json.loads(summary_path.read_text(encoding="utf-8"))
            state["pr_links"] = [args.pr_url]
            send_pipeline_report(state, run_id)
            print("[+] Pipeline report email sent successfully post-PR update.")
        else:
            print(f"[!] Summary file not found at {summary_path}. Cannot send post-PR email.")
    except Exception as e:
        print(f"[x] Error sending post-PR email: {e}")

if __name__ == "__main__":
    main()
